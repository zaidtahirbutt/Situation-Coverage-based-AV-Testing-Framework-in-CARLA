#!/usr/bin/env python

# Copyright (c) 2018-2020 Intel Corporation
#
# This work is licensed under the terms of the MIT license.
# For a copy, see <https://opensource.org/licenses/MIT>.

"""
Welcome to CARLA scenario_runner

This is the main script to be executed when running a scenario.
It loads the scenario configuration, loads the scenario and manager,
and finally triggers the scenario execution.
"""

from __future__ import print_function

import glob
import traceback
import argparse
from argparse import RawTextHelpFormatter
from datetime import datetime
from distutils.version import LooseVersion
import importlib
import inspect
import os
import signal
import sys
import time
import json
import pkg_resources
from random import seed
from random import randint
import scipy
import numpy as np
from scipy.special import softmax

import carla

from srunner.scenarioconfigs.openscenario_configuration import OpenScenarioConfiguration
from srunner.scenariomanager.carla_data_provider import CarlaDataProvider
from srunner.scenariomanager.scenario_manager11 import ScenarioManager  # 3 here now xD
from srunner.scenarios.open_scenario import OpenScenario
from srunner.scenarios.route_scenario import RouteScenario
from srunner.tools.scenario_parser import ScenarioConfigurationParser
from srunner.tools.route_parser import RouteParser
from automatic_control_agent_z5_other_veh import *  # OtherVehControlAgent() imported from here

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# Version of scenario_runner
VERSION = '0.9.9'


class ScenarioRunner(object):

    """
    This is the core scenario runner module. It is responsible for
    running (and repeating) a single scenario or a list of scenarios.

    Usage:
    scenario_runner = ScenarioRunner(args)
    scenario_runner.run()
    del scenario_runner
    """

    ego_vehicles = []

    # Tunable parameters
    client_timeout = 10.0  # in seconds
    #client_timeout = 30.0  # in seconds
    wait_for_world = 20.0  # in seconds
    #wait_for_world = 30.0  # in seconds
    frame_rate = 20.0      # in Hz

    # CARLA world and scenario handlers
    world = None
    manager = None

    additional_scenario_module = None

    agent_instance = None
    module_agent = None

    def __init__(self, args):
        """
        Setup CARLA client and world
        Setup ScenarioManager
        """
        self._args = args

        if args.timeout:
            self.client_timeout = float(args.timeout)

        # First of all, we need to create the client that will send the requests
        # to the simulator. Here we'll assume the simulator is accepting
        # requests in the localhost at port 2000.
        self.client = carla.Client(args.host, int(args.port))
        self.client.set_timeout(self.client_timeout)

        self.traffic_manager = self.client.get_trafficmanager(int(self._args.trafficManagerPort))

        dist = pkg_resources.get_distribution("carla")
        if LooseVersion(dist.version) < LooseVersion('0.9.8'):
            raise ImportError("CARLA version 0.9.8 or newer required. CARLA version found: {}".format(dist))

        # Load agent if requested via command line args
        # If something goes wrong an exception will be thrown by importlib (ok here)
        if self._args.agent is not None:  # yep not none
            module_name = os.path.basename(args.agent).split('.')[0]
            sys.path.insert(0, os.path.dirname(args.agent))
            self.module_agent = importlib.import_module(module_name)  # guess the agent class would be included here as library 

            #print(self.module_agent)  # <module 'npc_agent' from 'srunner/autoagents\\npc_agent.py'>
        
        # Create the ScenarioManager
        self.manager = ScenarioManager(self._args.debug, self._args.sync, self._args.timeout)

        # Create signal handler for SIGINT
        self._shutdown_requested = False
        if sys.platform != 'win32':
            signal.signal(signal.SIGHUP, self._signal_handler)
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)

        self._start_wall_time = datetime.now()

        self.fault_triggered = 0  # default value. 0 means no fault triggered.

    def destroy(self):
        """
        Cleanup and delete actors, ScenarioManager and CARLA world
        """

        self._cleanup()
        if self.manager is not None:
            del self.manager
        if self.world is not None:
            del self.world
        if self.client is not None:
            del self.client

    def _signal_handler(self, signum, frame):
        """
        Terminate scenario ticking when receiving a signal interrupt
        """
        self._shutdown_requested = True
        if self.manager:
            self.manager.stop_scenario()
            self._cleanup()
            if not self.manager.get_running_status():
                raise RuntimeError("Timeout occured during scenario execution")

    def _get_scenario_class_or_fail(self, scenario):  #  scenario = config.tpye = NoSignalJunctionCrossing
        """
        Get scenario class by scenario name
        If scenario is not supported or not found, exit script
        """

        # Path of all scenario at "srunner/scenarios" folder + the path of the additional scenario argument
        #The scenarios classes in the scenarios folder!!!! no_signal_junction_crossing.py is for our NoSignalJunctionCrossing xml file name
        scenarios_list = glob.glob("{}/srunner/scenarios/*.py".format(os.getenv('SCENARIO_RUNNER_ROOT', "./")))
        scenarios_list.append(self._args.additionalScenario)

        #print(scenarios_list)
        #return 0

        for scenario_file in scenarios_list:

            #one secnario_file is 'D:\\scenario_runner-0.9.10/srunner/scenarios\\no_signal_junction_crossing.py'
            # Get their module
            module_name = os.path.basename(scenario_file).split('.')[0]  # so base name would give us scenarios right?

            #print(module_name)

            ''' 
                These values were returned by the print command. So we got cut off at no sig junc meaning for loop broke there

                background_activity
                basic_scenario
                change_lane
                control_loss
                cut_in
                follow_leading_vehicle
                freeride
                junction_crossing_route
                maneuver_opposite_direction
                master_scenario
                no_signal_junction_crossing
                
            '''
            sys.path.insert(0, os.path.dirname(scenario_file))
            scenario_module = importlib.import_module(module_name)

            #print(scenario_module)


            '''
                These values were returned by the print command. So we got cut off at no sig junc meaning for loop broke there
                I think these scenario classes i.e., python files have been imported as libraries in this python scrip here
                pretty cool function importlib.import_module(module_name)

                <module 'background_activity' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\background_activity.py'>
                <module 'basic_scenario' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\basic_scenario.py'>
                <module 'change_lane' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\change_lane.py'>
                <module 'control_loss' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\control_loss.py'>
                <module 'cut_in' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\cut_in.py'>
                <module 'follow_leading_vehicle' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\follow_leading_vehicle.py'>
                <module 'freeride' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\freeride.py'>
                <module 'junction_crossing_route' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\junction_crossing_route.py'>
                <module 'maneuver_opposite_direction' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\maneuver_opposite_direction.py'>
                <module 'master_scenario' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\master_scenario.py'>
                <module 'no_signal_junction_crossing' from 'D:\\scenario_runner-0.9.10/srunner/scenarios\\no_signal_junction_crossing.py'> stops here cx we needed this scenario file and the loop breaks


            '''


            # And their members of type class

            '''
                
                The inspect module provides several useful functions to help get information about live objects such as modules, 
                classes, methods, functions, tracebacks, frame objects, and code objects. For example, it can help you examine 
                the contents of a class, retrieve the source code of a method, extract 
                and format the argument list for a function, or get all the information you need to display a detailed traceback
                
                inspect.getmembers(object[, predicate])¶
                Return all the members of an object in a list of (name, value) pairs sorted by name. If the optional predicate 
                argument—which will be called with the value object of 
                each member—is supplied, only members for which the predicate returns a true value are included.


                inspect.isclass(object)¶
                Return True if the object is a class, whether built-in or created in Python code.

            '''

            for member in inspect.getmembers(scenario_module, inspect.isclass):  # what's inspect doing?

                #So I got it. The name of the class of this no_signal_junction_crossing.py class is, "class NoSignalJunctionCrossing(BasicScenario):"!!!
                #Literally is NoSignalJunctionCrossing, the name by which we invoke the function. So it calls this class!!!!! 

                #print(member)  # heck lot of classes were printed here considering we went through all the iterations of
                # the for loop . e.g., ('WaypointFollower', <class 'srunner.scenariomanager.scenarioatomics.atomic_behaviors.WaypointFollower'>)
                #sys.exit('got em')  # ('BackgroundActivity', <class 'background_activity.BackgroundActivity'>)

                if scenario in member:  # scenario (i.e., class name) = NoSignalJunctionCrossing, for dictionary, if in statement looks at keys of the dictionary
                    
                    #print(member)  # ('NoSignalJunctionCrossing', <class 'no_signal_junction_crossing.NoSignalJunctionCrossing'>)
                    print(member[1])  # <class 'no_signal_junction_crossing.NoSignalJunctionCrossing'> was printed


                    return member[1]  # <class 'no_signal_junction_crossing.NoSignalJunctionCrossing'>

            # Remove unused Python paths
            sys.path.pop(0)

        print("Scenario '{}' not supported ... Exiting".format(scenario))
        sys.exit(-1)

    def _cleanup(self):
        """
        Remove and destroy all actors
        """
        # Simulation still running and in synchronous mode?
        if self.manager is not None and self.manager.get_running_status() \
                and self.world is not None and self._args.sync:
            # Reset to asynchronous mode
            settings = self.world.get_settings()
            settings.synchronous_mode = False
            settings.fixed_delta_seconds = None
            self.world.apply_settings(settings)

        self.manager.cleanup()

        CarlaDataProvider.cleanup()

        for i, _ in enumerate(self.ego_vehicles):
            if self.ego_vehicles[i]:
                if not self._args.waitForEgo:
                    print("Destroying ego vehicle {}".format(self.ego_vehicles[i].id))
                    self.ego_vehicles[i].destroy()
                self.ego_vehicles[i] = None
        self.ego_vehicles = []

        if self.agent_instance:
            self.agent_instance.destroy()
            self.agent_instance = None

    def _prepare_ego_vehicles(self, ego_vehicles):  # ego_vehicles is ActorConfigurationData(object):
        """
        Spawn or update the ego vehicles
        """

        #Enters here for the route scenario AS WELL.
        if not self._args.waitForEgo: # it's if NOT False, which is actually true damit! xD # if i have already defined an ego vehicle (not agent, agent is a script (brain) to control the vehicle)
            #print("HEEJCEJJCEJ")  # yep it was this one!
            #we are definitely passing rolename here but to this req new actor method
            #category="car" from sencario_configuration.py, def below
            '''
                def request_new_actor(model, spawn_point, rolename='scenario', autopilot=False,
                          random_location=False, color=None, actor_category="car"):
            
            '''
            for vehicle in ego_vehicles:  # empty for route scenario. Just checked, it doesn't have any ego vehicles in it's xml anyway xD
                self.ego_vehicles.append(CarlaDataProvider.request_new_actor(vehicle.model,
                                                                             vehicle.transform,
                                                                             vehicle.rolename, # ROLENAME IS HERO FOR EGO VEHICLES ZZZZZZZZZZZZ from scenario parser
                                                                             color=vehicle.color,
                                                                             actor_category=vehicle.category))

                #  We got our ego vehicle here and renamed it to Hero and spawned it at its spawn location at .transform
        #Doesn't enter here
        else:
            ego_vehicle_missing = True
            while ego_vehicle_missing:
                self.ego_vehicles = []
                ego_vehicle_missing = False
                for ego_vehicle in ego_vehicles:
                    ego_vehicle_found = False
                    carla_vehicles = CarlaDataProvider.get_world().get_actors().filter('vehicle.*')
                    for carla_vehicle in carla_vehicles:

                        # doesn't enter this loop either. wt? Cx ego_vehicle_missing isn't initialized
                        # print("HEEJCEJJCEJ")
                        # print(carla_vehicle.attributes['role_name'])
                        # print(carla_vehicle.attributes['number_of_wheels'])
                        # print(carla_vehicle.attributes[ego_vehicle.rolename])
                        # return 0    

                        if carla_vehicle.attributes['role_name'] == ego_vehicle.rolename:
                            
                            # doesn't enter this loop either. Leme look at these freakin role names man! Does any of these have a role name 'hero'?
                            # print("HEEJCEJJCEJ")
                            # print(carla_vehicle.attributes['role_name'])  # print the value of index named 'role_name', probably is a dictionary
                            # print(carla_vehicle.attributes['number_of_wheels'])
                            # print(carla_vehicle.attributes[ego_vehicle.rolename])
                            # return 0



                            ego_vehicle_found = True
                            self.ego_vehicles.append(carla_vehicle)  # the empty ego_vehicles list at the start of the class above the init method
                            break
                    if not ego_vehicle_found:
                        ego_vehicle_missing = True
                        break
            #Doesn't enter this anyway for scenarios or route scenarios
            for i, _ in enumerate(self.ego_vehicles):
                self.ego_vehicles[i].set_transform(ego_vehicles[i].transform)  # Teleports the actor to a given transform (location and rotation).
                CarlaDataProvider.register_actor(self.ego_vehicles[i])

        # sync state
        if CarlaDataProvider.is_sync_mode():
            self.world.tick()
        else:
            self.world.wait_for_tick()

    def _analyze_scenario(self, config):
        """
        Provide feedback about success/failure of a scenario
        """

        # Create the filename
        current_time = str(datetime.now().strftime('%Y-%m-%d-%H-%M-%S'))
        junit_filename = None
        config_name = config.name
        if self._args.outputDir != '':
            config_name = os.path.join(self._args.outputDir, config_name)
        if self._args.junit:
            junit_filename = config_name + current_time + ".xml"
        filename = None
        if self._args.file:
            filename = config_name + current_time + ".txt"

        if not self.manager.analyze_scenario(self._args.output, filename, junit_filename):
            print("All scenario tests were passed successfully!")
        else:
            print("Not all scenario tests were successful")
            if not (self._args.output or filename or junit_filename):
                print("Please run with --output for further information")

    def _record_criteria(self, criteria, name):
        """
        Filter the JSON serializable attributes of the criterias and
        dumps them into a file. This will be used by the metrics manager,
        in case the user wants specific information about the criterias.
        """
        file_name = name[:-4] + ".json"

        # Filter the attributes that aren't JSON serializable
        with open('temp.json', 'w') as fp:

            criteria_dict = {}
            for criterion in criteria:

                criterion_dict = criterion.__dict__
                criteria_dict[criterion.name] = {}

                for key in criterion_dict:
                    if key != "name":
                        try:
                            key_dict = {key: criterion_dict[key]}
                            json.dump(key_dict, fp, sort_keys=False, indent=4)
                            criteria_dict[criterion.name].update(key_dict)
                        except TypeError:
                            pass

        os.remove('temp.json')

        # Save the criteria dictionary into a .json file
        with open(file_name, 'w') as fp:
            json.dump(criteria_dict, fp, sort_keys=False, indent=4)

    def _load_and_wait_for_world(self, town, ego_vehicles=None):
        """
        Load a new CARLA world and provide data to CarlaDataProvider
        """

        if self._args.reloadWorld:
            self.world = self.client.load_world(town)  # particular town loaded in the world
        else:
            # if the world should not be reloaded, wait at least until all ego vehicles are ready
            ego_vehicle_found = False
            if self._args.waitForEgo:  # not for routes either
                while not ego_vehicle_found and not self._shutdown_requested:
                    vehicles = self.client.get_world().get_actors().filter('vehicle.*')
                    for ego_vehicle in ego_vehicles:
                        ego_vehicle_found = False
                        for vehicle in vehicles:
                            if vehicle.attributes['role_name'] == ego_vehicle.rolename:
                                
                                # doesn't enter this loop cx args.wait for ego is false
                                # print("HEEJCEJJCEJ")
                                # print(vehicle.attributes['role_name'])
                                # print(vehicle.attributes['number_of_wheels'])
                                # return 0
                                
                                ego_vehicle_found = True
                                break
                        if not ego_vehicle_found:
                            print("Not all ego vehicles ready. Waiting ... ")
                            time.sleep(1)
                            break

        self.world = self.client.get_world()

        if self._args.sync:
            settings = self.world.get_settings()  # World settings
            settings.synchronous_mode = True
            settings.fixed_delta_seconds = 1.0 / self.frame_rate
            self.world.apply_settings(settings)

        self.traffic_manager.set_synchronous_mode(True)
        self.traffic_manager.set_random_device_seed(int(self._args.trafficManagerSeed))

        CarlaDataProvider.set_client(self.client)
        CarlaDataProvider.set_world(self.world)
        CarlaDataProvider.set_traffic_manager_port(int(self._args.trafficManagerPort))

        # Wait for the world to be ready
        if CarlaDataProvider.is_sync_mode():
            self.world.tick()
        else:
            self.world.wait_for_tick()
        if CarlaDataProvider.get_map().name != town and CarlaDataProvider.get_map().name != "OpenDriveMap":
            print("The CARLA server uses the wrong map: {}".format(CarlaDataProvider.get_map().name))
            print("This scenario requires to use map: {}".format(town))
            return False

        #So in this function we're only loading the map and making the client-server connection, by the default choices we
        #aren't loading the ego vehicle from Carla API yet!

        return True

    def _load_and_run_scenario(self, config, intersection_situations):  # config is ScenarioConfiguration(object): from scenario parser   # Config has all parameters like ego and other veh (ActorConfigurationData(object): objects), their details
        # like locations vheicle transforms and types, weather, towns, trigger points (ego veh spawn loc), routes if any etc 
        #BasicScenario class
            #if self.config.friction is not None: we can introduce friction triggers in conig file!
            # See the use of this friction in basic_scenario. wait leme see :3
        """
        Load and run the scenario given by config
        """
        result = False
        if not self._load_and_wait_for_world(config.town, config.ego_vehicles):  # so we got the world from this!  # config.ego_vehicles is empty for routes scenario
            self._cleanup()
            return False

        if self._args.agent:  # only compatable with route based scenarios that have agents! #print(self.module_agent)  # <module 'npc_agent' from 'srunner/autoagents\\npc_agent.py'>
            agent_class_name = self.module_agent.__name__.title().replace('_', '')
            #agent_class_name2 = self.module_agent.__name__.title()
            #print(self.module_agent)  # <module 'npc_agent' from 'srunner/autoagents\\npc_agent.py'>
            #print(agent_class_name)  # NpcAgent
            #print(agent_class_name2)  # Npc_Agent

            try:
                self.agent_instance = getattr(self.module_agent, agent_class_name)(self._args.agentConfig)
                #Return the value of the named attribute of object. name must be a string.

                #print(self.agent_instance)  # <npc_agent.NpcAgent object at 0x000001F4E79A6088>

                config.agent = self.agent_instance  # <npc_agent.NpcAgent object at 0x000001F4E79A6088>. Config from routes parser.
            except Exception as e:          # pylint: disable=broad-except
                traceback.print_exc()
                print("Could not setup required agent due to {}".format(e))
                self._cleanup()
                return False  # Will learn about this when I work with agents IA

        # Prepare scenario
        print("Preparing scenario: " + config.name)
        try:
            
            #ego vehicles empty right now for routes scenario
            
            self._prepare_ego_vehicles(config.ego_vehicles)  # imp method # ROLENAME IS HERO FOR EGO VEHICLES ZZZZZZZZZZZZ from scenario_parser.py
            
            #Nothing in here for routes: I'm thinking we don't need an ego vehicle? Cx we want our npc's to wreck havoc anyway. So maybe we don't need other vehicles either also?
            #cx we can just initialize all npcs inside the Routes Scenario class? .. We'll see.

            #So we have other vehicles transforms and their front, left and right positions and town names and just transforms inside the config.scenario, i.e., the scenario json file. Let's see what it's use is. 




            if self._args.openscenario:
                scenario = OpenScenario(world=self.world,
                                        ego_vehicles=self.ego_vehicles,
                                        config=config,
                                        config_file=self._args.openscenario,
                                        timeout=100000)
            elif self._args.route:
                scenario = RouteScenario(world=self.world,
                                         config=config,
                                         debug_mode=self._args.debug)
            else:
                scenario_class = self._get_scenario_class_or_fail(config.type)  #  config.tpye = NoSignalJunctionCrossing,  <class 'no_signal_junction_crossing.NoSignalJunctionCrossing'>

                #print(member[1])  # <class 'no_signal_junction_crossing.NoSignalJunctionCrossing'> was returned in scenario class
                #i.e., NoSignalJunctionCrossing class (address maybe, as an imported library/module) was returned to this variable scenario_class
                #because we are initializing NoSignalJunctionCrossing class by initializing scenario_class() below

                # Config has all parameters like ego and other veh, their details (taken from xml files in examples folder for their configuartions using scenario_parser and scenario_configuration.py)
                # like locations vheicle transforms and types, weather, towns, trigger points (ego veh spawn loc), routes if any etc
                
                #BasicScenario class
                    #if self.config.friction is not None: we can introduce friction triggers in conig file!

                    # Initializing adversarial actors
                    # self._initialize_actors(config)
                        #Default initialization of other actors.
                        #Override this method in child class to provide custom initialization.
                #Initialized here
                    #The lengthy initialization of Basic scenario also happens here inside the NoSignalJunctionCrossing class inside it's
                    #super__init__ function inside it's __init__ functions 
                
                # These are _args.show_int_wp & _args.show_int_wp_locs are acting as gatekeepers to the new edited IntersectionScenario class xD
                
                
                ################################ Initialize Other Vehicle Agent here!!!!!!!!!!!!! ##############################
                self.other_veh_agentZ_ogg = OtherVehControlAgent()




                #if self._args.show_int_wps or self._args.show_int_wp_locs:  # So when we call either of these arguments, that means we are calling our custom scenario class!
                self.xyz = 1
                if self.xyz == 1:
                # ***************************************** IntersectionSituations class here **********************    

                # After all parameters are decided in the block of code above, we pass that to the new scenario config file or wait, maybe we can use the above block of code in the method before the load and run scenario. Wait leme check.
                # 

                    scenario = scenario_class(self.world,
                                              self.ego_vehicles,
                                              config,
                                              intersection_situations,
                                              self.other_veh_agentZ_ogg,
                                              randomize=self._args.randomize,
                                              debug_mode=self._args.debug,
                                              show_int_wps=self._args.show_int_wps,
                                              show_int_wp_locs=self._args.show_int_wp_locs)  # so without calling by keyword, we are using position arguments to a function and hence self._args.show_int_wp_locs here = criteria_enable 

                    #print("entered here")  # yesss
                    #print(self._args.show_int_wp_locs)  # True when --show_int_wp_locs is typed and False when it is not typed
                    #sys.exit("offer")

                else:
                    pass
                    # scenario = scenario_class(self.world,
                    #                           self.ego_vehicles,
                    #                           config,
                    #                           self._args.randomize,
                    #                           self._args.debug)
                    
                    #print("entered here as well")  # noo
                    #sys.exit("offer 2")
        
        except Exception as exception:                  # pylint: disable=broad-except
            print("The scenario cannot be loaded")
            traceback.print_exc()
            print(exception)
            self._cleanup()
            return False

        try:  # try catch, great for debugging
            if self._args.record:  # recordings of the scenario run maybe
                recorder_name = "{}/{}/{}.log".format(
                    os.getenv('SCENARIO_RUNNER_ROOT', "./"), self._args.record, config.name)
                self.client.start_recorder(recorder_name, True)


            # Calculating ego destination and passing it off to scenario manager load_scenario method

            self.ego_goal_loc_x = intersection_situations.ego_veh.ego_endconditiontrigger_dict["key_ego_destination_x"]
            self.ego_goal_loc_y = intersection_situations.ego_veh.ego_endconditiontrigger_dict["key_ego_destination_y"]

            self.ego_goal_carla_Location = carla.Location(self.ego_goal_loc_x, self.ego_goal_loc_y, 0)

            #print("this is the ego goal location", self.ego_goal_carla_Location)

            #self.visualize = True  # for visualizing the scenario
            self.visualize = self._args.not_visualize  # for visualizing the scenario. self._args.not_visualize is true as default, so we VISUALIZE on default if this argument is not mentioned in the main script call

            #print("vis", self.visualize)
            #sys.exit("hawk")

            # Load scenario and run it  # IMPORTANT *********************************S
            self.manager.load_scenario(scenario, self.ego_goal_carla_Location, self.other_veh_agentZ_ogg, self.visualize, self.agent_instance)  # damn big function/method  # self.agent_instance is None  for non route based scenarios
            self.manager.run_scenario()

            # Provide outputs if required
            self._analyze_scenario(config)  # will look at this at a later stage when I get custom scenarios running iA


            

            #self.write_results_to_excel_file()  # using these self.manager.collision_counts, self.manager.collision_test_result, self.counting_reps, self.fault_triggered


            # Remove all actors, stop the recorder and save all criterias (if needed)
            scenario.remove_all_actors()
            if self._args.record:
                self.client.stop_recorder()
                self._record_criteria(self.manager.scenario.get_criteria(), recorder_name)

            result = True

        except Exception as e:              # pylint: disable=broad-except
            traceback.print_exc()
            print(e)
            result = False

        self._cleanup()
        return result

    def _run_scenarios(self):
        """
        Run conventional scenarios (e.g. implemented using the Python API of ScenarioRunner)
        """
        result = False

        # Load the scenario configurations provided in the config file #  arg.scenario e.g. is NoSignalJunction
        # print("Hello1")
        # print(self._args.scenario)  # NoSignalJunctionCrossing
        # print("Hello2")
        # print(self._args.configFile)  # Config file is '' for NoSignalJunction
        # print("Hello3")
        # return 0

        scenario_configurations = ScenarioConfigurationParser.parse_scenario_configuration(
            self._args.scenario,
            self._args.configFile)  # ScenarioConfiguration(object): inside list of scenario_configurations
        if not scenario_configurations:
            print("Configuration for scenario {} cannot be found!".format(self._args.scenario))
            return result

        #print("HELLO")
        #print(scenario_configurations[0].ego_vehicles)
        #print(scenario_configurations[0].ego_vehicles[0])
        #print("HELLO2")
        #return 0

        # Flag to start generating environmental conditions
        #self.activate_env_cond_generation = True #True #False
        self.activate_env_cond_generation = self._args.disable_env_cond_gen  #default value is true unless I call this argument in main file call        
        try:    
            # Initialize the Sit Generation object
            #self.intersection_situations = IntersectionSituations()
            #self.intersection_situations = IntersectionSituations(self._args.Activate_IntersectionScenario_Seed, self._args.IntersectionScenario_Seed, self._args.use_sit_cov)  # default value of self._args.use_sit_cov is False!!!!!
            self.intersection_situations = IntersectionSituations(self._args.Activate_IntersectionScenario_Seed, self._args.IntersectionScenario_Seed)  # default value of self._args.use_sit_cov is False!!!!!
        except Exception as e: 
            traceback.print_exc()
            print("Could not setup IntersectionSituations due to {}".format(e))
            self._cleanup()
            return False 

        self.counting_reps = 0 
        # Execute each configuration
        for config in scenario_configurations:
            
            # ************************************** THIS IS THE LOOP THAT CONTROLS repititions!!!!!!!! #####################################################
            
            for _ in range(self._args.repetitions):  #  put 2 in this and it repeated from 00! # lets see what happens in reptitions
                
                # default repetitions = 1 :3

                #So config is the iterator through the list of different scenario configurations in the scenario_configurations list 
                ''' 
                the last lines from parser function of scenario config, just for clarity 

                    scenario_configurations.append(new_config)

                return scenario_configurations

                '''
                #print("HELLO")
                #print(config.name)  # NoSignalJunctionCrossing was the output 
                #print(config.type)  # NoSignalJunctionCrossing was the output
                #print(config.town)  # Town03 was the output 
                #return 0


                # IntersectionScenario class here??? Then pass the config arguments to the new scenario parser file? But what about the parameters that I want to pass to the init method of the scenario directly? Pass them through the load and run scenario :3 .... As an object probably :3, Would that be too heavy?
                # I don't think just passing some objects containing small data would be heavy. List down what we need to pass and then pass the objects in the init method of load and run,
                # Initialize IntersectionScenario object here. So here we have all the access. Pass them to the method to retain access :3 But if the main objective is to edit the config file here, I guess that would make things a lot neater and will have the proper flow.

                # self.intersection_situations = IntersectionSituations("random")
                try:
                    # here we select sit generation parameters RANDOMLY     
                    #self.intersection_situations.start_sit_config_gen("random", self.activate_env_cond_generation)
                    if self._args.use_sit_cov == False:
                        approach = "random"
                    else:
                        approach = "sitcov"

                    #self.intersection_situations.start_sit_config_gen("random", self.activate_env_cond_generation)
                    self.intersection_situations.start_sit_config_gen(approach, self.activate_env_cond_generation)

                    # Just checking if the code runs. Comment out below when not debugging. 
                    
                    # ego_veh_transform = self.intersection_situations.ego_veh.ego_start_carla_transform_dict["key_transform"]
                    # print(ego_veh_transform.location)
                    # print("Location from top is",self.intersection_situations.ego_veh.ego_start_carla_transform_dict["key_location_string"])
                    # sys.exit("Sys.exit at _load_and_run_scenario, main script. IntersectionSituations executed successfully")
                except Exception as e: 
                    traceback.print_exc()
                    print("Could not setup IntersectionSituations due to {}".format(e))
                    self._cleanup()
                    return False              

                # Pass parameters to config file below!! :3
                # So we have the config file which is a ScenarioConfiguration() object.
                # ego_vehicles inside ScenarioConfiguration() object is a list. So "for vehicle in ego_vehicles:"" is done to do requestactor from carla data manger.
                # I can do ego_vehicle = ego_vehicles[-1]. And then I have the ActorConfigurationData(object): of the ego vehicle.
                # I can then do self.transform = transform

                # Ego Vehicle ActorConfigurationData(object):
                #self.ego_vehicle_actor_config_object = config.ego_vehicles[-1]

                # Overwrite the transform of the ego vehicle ActorConfigObject isnide the config file instead of making a new copy of it as above
                config.ego_vehicles[-1].transform = self.intersection_situations.ego_veh.ego_start_carla_transform_dict["key_transform"]
                
                print(config.ego_vehicles[-1].transform)
                print("Ego Veh Location from top is",self.intersection_situations.ego_veh.ego_start_carla_transform_dict["key_location_string"])
                # sys.exit("Sys.exit at _load_and_run_scenario, main script. IntersectionSituations executed successfully")  # Successfull!

                # Overwrite the transform of the Other vehicle ActorConfigObject isnide the config file
                config.other_actors[-1].transform = self.intersection_situations.other_veh.other_vehicle_start_carla_transform_dict["key_transform"]

                print(config.other_actors[-1].transform)
                print("Other Veh Location from top is",self.intersection_situations.other_veh.other_vehicle_start_carla_transform_dict["key_location_string"])
                # sys.exit("Sys.exit at _load_and_run_scenario, main script. IntersectionSituations executed successfully")


                # Overwrite the transform of ego veh location in config.triggerpoints list
                # Using 0 for the index because in basic_scenario.py zero is used for the index there
                config.trigger_points[0] = config.ego_vehicles[-1].transform  # This came from above

                # Overwriting config environemntal conditions values
                config.weather.cloudiness = self.intersection_situations.env_conditions.cloudiness
                config.weather.precipitation = self.intersection_situations.env_conditions.precipitation
                config.weather.precipitation_deposits = self.intersection_situations.env_conditions.precipitation_deposits
                config.weather.wind_intensity = self.intersection_situations.env_conditions.wind_intensity
                config.weather.sun_azimuth_angle = self.intersection_situations.env_conditions.sun_azimuth_angle
                config.weather.sun_altitude_angle = self.intersection_situations.env_conditions.sun_altitude_angle
                config.weather.fog_density = self.intersection_situations.env_conditions.fog_density
                config.weather.fog_distance = self.intersection_situations.env_conditions.fog_distance
                config.weather.wetness = self.intersection_situations.env_conditions.wetness
                config.weather.friction = self.intersection_situations.env_conditions.friction

                print("friction is", config.weather.friction)
                #sys.exit("block")

                # This parameter isn't included yet in the config class. Edit and test it later
                #config.weather.fog_falloff = self.intersection_situations.env_conditions.fog_falloff








                # Ego Vehicle and Other Vehicle transforms have been overwritten successfully.
                # Now to pass on the Different trigger locations to IntersectionScenarioZ_# class
                # To do that I will pass on the IntersectionSiuations object to the load and run scenario and from there I will pass it on to IntersectionScenarioZ_# class initialization iA :3

                result = self._load_and_run_scenario(config, self.intersection_situations)
                self.counting_reps = self.counting_reps + 1
                print(f"````````` Counter Reps = {self.counting_reps}`````````````")

            self._cleanup()
        return result  # reutrned True if no Exception else False.

    def _run_route(self):
        """
        Run the route scenario
        """
        result = False

        if self._args.route:
            routes = self._args.route[0]  # srunner/data/routes_debug.xml
            scenario_file = self._args.route[1]  # srunner/data/all_towns_traffic_scenarios1_3_4.json
            single_route = None
            if len(self._args.route) > 2:
                single_route = self._args.route[2]  # route id = 0; inside routes_debug.xml (input: (route_file,scenario_file,[route id]))

        # retrieve routes
        route_configurations = RouteParser.parse_routes_file(routes, scenario_file, single_route)  # return list_route_descriptions  -> list_route_descriptions.append(new_config)  # only one route cx of one route ID. This list has only one object hence. --> new_config = RouteScenarioConfiguration()  # RouteScenarioConfiguration/scenario config object

        # Setting the scenario configuartions like in the run scenario method. But also parsing the routes along with weathers and ego vehicles and other vehicles from xml file hawks.



        for config in route_configurations:
            for _ in range(self._args.repetitions):
                result = self._load_and_run_scenario(config)  # only weather and waypoints in this config right now along with -> new_config.name = "RouteScenario_{}".format(route_id)  # RouteScenario_0 and new_config.scenario_file = scenario_file  # The JSON file -> scenario_file -> srunner/data/all_towns_traffic_scenarios1_3_4.json

                self._cleanup()
        return result

    def _run_openscenario(self):
        """
        Run a scenario based on OpenSCENARIO
        """

        # Load the scenario configurations provided in the config file
        if not os.path.isfile(self._args.openscenario):
            print("File does not exist")
            self._cleanup()
            return False

        config = OpenScenarioConfiguration(self._args.openscenario, self.client)

        result = self._load_and_run_scenario(config)
        self._cleanup()
        return result

    def run(self):
        """
        Run all scenarios according to provided commandline args
        """
        result = True
        if self._args.openscenario:
            result = self._run_openscenario()
        elif self._args.route:  # now we enter here for routes :3
            result = self._run_route()
        else:
            result = self._run_scenarios()

        print("No more scenarios .... Exiting")
        return result

    def write_results_to_excel_file(self):   # using these self.manager.collision_counts, self.manager.collision_test_result, self.counting_reps, self.fault_triggered

        dest_filename = 'SituationCoverage_AVTesting_Framework_results.xlsx'  # convert this as argument maybe? or no :3

        cell_font = Font(name='Calibri', size=28, bold=True)
        cell_smaller_font = Font(name='Calibri', size=22, bold=True)
        cell_alignment = Alignment(horizontal="center", vertical="center")

        wb = load_workbook(filename=dest_filename)
        
        Sr_no_column = 1

        pad_rows = 0 + 20 + 20 + 20 + 20 # this is for if simulation breaks before all repititions are completed

        if self.manager.collision_counts > 0:
            collision = 1
        else:
            collision = 0

        fault_1_triggered = 0
        fault_2_triggered = 1
        fault_3_triggered = 0
        no_fault_triggered = 0

        if self.fault_triggered == 0:
            no_fault_triggered = 1
        elif self.fault_triggered == 1:
            fault_1_triggered = 1
        elif self.fault_triggered == 2:
            fault_2_triggered = 1
        elif self.fault_triggered == 3:
            fault_3_triggered = 1
        else:
            raise ValueError("Select Appropriate Fault")
        


        # for env conditions sheet
        #sheet = wb['Sheet1env']
        sheet = wb['Sheet1env_fault2']
        sheet2 = wb['Sheet2int_fault2']

        env_starting_row = 4 + pad_rows
        cloudiness_dict_start_column_excel = 2
        precipitation_dict_start_column_excel = 8
        precipitation_deposits_dict_start_column_excel = 14
        wind_intensity_dict_start_column_excel = 20 
        sun_azimuth_angle_dict_start_column_excel = 26
        sun_altitude_angle_dict_start_column_excel = 33
        fog_density_dict_start_column_excel = 40
        fog_distance_dict_start_column_excel = 46
        wetness_dict_start_column_excel = 52
        fog_falloff_dict_start_column_excel = 58
        friction_dict_start_column_excel = 64

        env_collision_1_or_0_column = 70
        env_collision_counts_column = 71
        env_fault_1_triggered_column = 72
        env_fault_2_triggered_column = 73
        env_fault_3_triggered_column = 74
        env_no_fault_triggered_column = 75  
        env_collision_test_result_column = 76

        env_current_row = env_starting_row + self.counting_reps

        sheet.cell(env_current_row, Sr_no_column).value = 1 + self.counting_reps
        sheet.cell(env_current_row, Sr_no_column).font = cell_font
        sheet.cell(env_current_row, Sr_no_column).alignment = cell_alignment

        # The dictionaries used for env conditions
        # self.cloudiness_key
        # self.precipitation_key
        # self.precipitation_deposits_key
        # self.wind_intensity_key
        # self.fog_density_key
        # self.fog_distance_key 
        # self.wetness_key
        # self.fog_falloff_key
        # self.friction_key

        # self.cloudiness_key
        sheet.cell(env_current_row, cloudiness_dict_start_column_excel + self.intersection_situations.cloudiness_key).value = 1
        sheet.cell(env_current_row, cloudiness_dict_start_column_excel + self.intersection_situations.cloudiness_key).font = cell_font
        sheet.cell(env_current_row, cloudiness_dict_start_column_excel + self.intersection_situations.cloudiness_key).alignment = cell_alignment

        # self.precipitation_key
        sheet.cell(env_current_row, precipitation_dict_start_column_excel + self.intersection_situations.precipitation_key).value = 1
        sheet.cell(env_current_row, precipitation_dict_start_column_excel + self.intersection_situations.precipitation_key).font = cell_font
        sheet.cell(env_current_row, precipitation_dict_start_column_excel + self.intersection_situations.precipitation_key).alignment = cell_alignment

        # self.precipitation_deposits_key
        sheet.cell(env_current_row, precipitation_deposits_dict_start_column_excel + self.intersection_situations.precipitation_deposits_key).value = 1
        sheet.cell(env_current_row, precipitation_deposits_dict_start_column_excel + self.intersection_situations.precipitation_deposits_key).font = cell_font
        sheet.cell(env_current_row, precipitation_deposits_dict_start_column_excel + self.intersection_situations.precipitation_deposits_key).alignment = cell_alignment

        # self.wind_intensity_key
        sheet.cell(env_current_row, wind_intensity_dict_start_column_excel + self.intersection_situations.wind_intensity_key).value = 1
        sheet.cell(env_current_row, wind_intensity_dict_start_column_excel + self.intersection_situations.wind_intensity_key).font = cell_font
        sheet.cell(env_current_row, wind_intensity_dict_start_column_excel + self.intersection_situations.wind_intensity_key).alignment = cell_alignment

        # self.fog_density_key
        sheet.cell(env_current_row, fog_density_dict_start_column_excel + self.intersection_situations.fog_density_key).value = 1
        sheet.cell(env_current_row, fog_density_dict_start_column_excel + self.intersection_situations.fog_density_key).font = cell_font
        sheet.cell(env_current_row, fog_density_dict_start_column_excel + self.intersection_situations.fog_density_key).alignment = cell_alignment

        # self.fog_distance_key
        sheet.cell(env_current_row, fog_distance_dict_start_column_excel + self.intersection_situations.fog_distance_key).value = 1
        sheet.cell(env_current_row, fog_distance_dict_start_column_excel + self.intersection_situations.fog_distance_key).font = cell_font
        sheet.cell(env_current_row, fog_distance_dict_start_column_excel + self.intersection_situations.fog_distance_key).alignment = cell_alignment

        # self.wetness_key
        sheet.cell(env_current_row, wetness_dict_start_column_excel + self.intersection_situations.wetness_key).value = 1
        sheet.cell(env_current_row, wetness_dict_start_column_excel + self.intersection_situations.wetness_key).font = cell_font
        sheet.cell(env_current_row, wetness_dict_start_column_excel + self.intersection_situations.wetness_key).alignment = cell_alignment

        # self.fog_falloff_key
        sheet.cell(env_current_row, fog_falloff_dict_start_column_excel + self.intersection_situations.fog_falloff_key).value = 1
        sheet.cell(env_current_row, fog_falloff_dict_start_column_excel + self.intersection_situations.fog_falloff_key).font = cell_font
        sheet.cell(env_current_row, fog_falloff_dict_start_column_excel + self.intersection_situations.fog_falloff_key).alignment = cell_alignment

        # self.friction_key
        sheet.cell(env_current_row, friction_dict_start_column_excel + self.intersection_situations.friction_key).value = 1
        sheet.cell(env_current_row, friction_dict_start_column_excel + self.intersection_situations.friction_key).font = cell_font
        sheet.cell(env_current_row, friction_dict_start_column_excel + self.intersection_situations.friction_key).alignment = cell_alignment


        # ENV RESULTS

        # env_collision_1_or_0_column 
        sheet.cell(env_current_row, env_collision_1_or_0_column).value = collision
        sheet.cell(env_current_row, env_collision_1_or_0_column).font = cell_font
        sheet.cell(env_current_row, env_collision_1_or_0_column).alignment = cell_alignment

        # env_collision_counts_column
        sheet.cell(env_current_row, env_collision_counts_column).value = self.manager.collision_counts
        sheet.cell(env_current_row, env_collision_counts_column).font = cell_font
        sheet.cell(env_current_row, env_collision_counts_column).alignment = cell_alignment

        # env_fault_1_triggered_column
        sheet.cell(env_current_row, env_fault_1_triggered_column).value = fault_1_triggered
        sheet.cell(env_current_row, env_fault_1_triggered_column).font = cell_font
        sheet.cell(env_current_row, env_fault_1_triggered_column).alignment = cell_alignment

        # env_fault_2_triggered_column
        sheet.cell(env_current_row, env_fault_2_triggered_column).value = fault_2_triggered
        sheet.cell(env_current_row, env_fault_2_triggered_column).font = cell_font
        sheet.cell(env_current_row, env_fault_2_triggered_column).alignment = cell_alignment

        # env_fault_3_triggered_column
        sheet.cell(env_current_row, env_fault_3_triggered_column).value = fault_3_triggered
        sheet.cell(env_current_row, env_fault_3_triggered_column).font = cell_font
        sheet.cell(env_current_row, env_fault_3_triggered_column).alignment = cell_alignment

        # env_no_fault_triggered_column
        sheet.cell(env_current_row, env_no_fault_triggered_column).value = no_fault_triggered
        sheet.cell(env_current_row, env_no_fault_triggered_column).font = cell_font
        sheet.cell(env_current_row, env_no_fault_triggered_column).alignment = cell_alignment

        # env_collision_test_result_column

        sheet.cell(env_current_row, env_collision_test_result_column).value = self.manager.collision_test_result
        sheet.cell(env_current_row, env_collision_test_result_column).font = cell_font
        sheet.cell(env_current_row, env_collision_test_result_column).alignment = cell_alignment





        # for intersections sheet
        #sheet2 = wb['Sheet2int']
        #sheet2 = wb['Sheet2int (2)']

        int_starting_row = 5 + pad_rows

        int_current_row = int_starting_row + self.counting_reps

        int_collision_1_or_0_column = 14
        int_collision_counts_column = 15
        int_fault_1_triggered_column = 16
        int_fault_2_triggered_column = 17
        int_fault_3_triggered_column = 18
        int_no_fault_triggered_column = 19
        int_collision_test_result_column = 20



        sheet2.cell(int_current_row, Sr_no_column).value = 1 + self.counting_reps
        sheet2.cell(int_current_row, Sr_no_column).font = cell_font
        sheet2.cell(int_current_row, Sr_no_column).alignment = cell_alignment

        key = self.intersection_situations.key_ego_other_veh_interaction_key

        if key == "key_SAVL_GBAVxSOVR_GBOV":
            int_sheet_column = 2
        elif key == "key_SAVL_GRAVxSOVR_GBOV":
            int_sheet_column = 3
        elif key == "key_SAVL_GRAVxSOVB_GROV":
            int_sheet_column = 4
        elif key == "key_SAVL_GRAVxSOVB_GLOV":
            int_sheet_column = 5
        elif key == "key_SAVB_GLAV_SOVL_GROV":
            int_sheet_column = 6
        elif key == "key_SAVB_GLAV_SOVR_GLOV":
            int_sheet_column = 7
        elif key == "key_SAVB_GLAVxSOVR_GBOV":
            int_sheet_column = 8
        elif key == "key_SAVB_GRAV_SOVL_GROV":
            int_sheet_column = 9
        elif key == "key_SAVR_GLAVxSOVB_GLOV":
            int_sheet_column = 10
        elif key == "key_SAVR_GBAVxSOVB_GLOV":
            int_sheet_column = 11
        elif key == "key_SAVR_GBAVxSOVL_GROV":
            int_sheet_column = 12
        elif key == "key_SAVR_GBAV_SOVL_GBOV":
            int_sheet_column = 13
        else:
            raise ValueError("Unexpected value of key")

        sheet2.cell(int_current_row, int_sheet_column).value = 1
        sheet2.cell(int_current_row, int_sheet_column).font = cell_font
        sheet2.cell(int_current_row, int_sheet_column).alignment = cell_alignment


        # Int RESULTS

        # int_collision_1_or_0_column 
        sheet2.cell(int_current_row, int_collision_1_or_0_column).value = collision
        sheet2.cell(int_current_row, int_collision_1_or_0_column).font = cell_font
        sheet2.cell(int_current_row, int_collision_1_or_0_column).alignment = cell_alignment

        # int_collision_counts_column
        sheet2.cell(int_current_row, int_collision_counts_column).value = self.manager.collision_counts
        sheet2.cell(int_current_row, int_collision_counts_column).font = cell_font
        sheet2.cell(int_current_row, int_collision_counts_column).alignment = cell_alignment

        # int_fault_1_triggered_column
        sheet2.cell(int_current_row, int_fault_1_triggered_column).value = fault_1_triggered
        sheet2.cell(int_current_row, int_fault_1_triggered_column).font = cell_font
        sheet2.cell(int_current_row, int_fault_1_triggered_column).alignment = cell_alignment

        # int_fault_2_triggered_column
        sheet2.cell(int_current_row, int_fault_2_triggered_column).value = fault_2_triggered
        sheet2.cell(int_current_row, int_fault_2_triggered_column).font = cell_font
        sheet2.cell(int_current_row, int_fault_2_triggered_column).alignment = cell_alignment

        # int_fault_3_triggered_column
        sheet2.cell(int_current_row, int_fault_3_triggered_column).value = fault_3_triggered
        sheet2.cell(int_current_row, int_fault_3_triggered_column).font = cell_font
        sheet2.cell(int_current_row, int_fault_3_triggered_column).alignment = cell_alignment

        # int_no_fault_triggered_column
        sheet2.cell(int_current_row, int_no_fault_triggered_column).value = no_fault_triggered
        sheet2.cell(int_current_row, int_no_fault_triggered_column).font = cell_font
        sheet2.cell(int_current_row, int_no_fault_triggered_column).alignment = cell_alignment

        # int_collision_test_result_column

        sheet2.cell(int_current_row, int_collision_test_result_column).value = self.manager.collision_test_result
        sheet2.cell(int_current_row, int_collision_test_result_column).font = cell_font
        sheet2.cell(int_current_row, int_collision_test_result_column).alignment = cell_alignment

        wb.save(filename = dest_filename)



class IntersectionSituations:
# SituationGeneration Engine!
    #def __init__(self, approach="random"):  # approach can either be "random" or "sitcov"

    # self.key_ego_other_veh_interaction_key is for outputting the key used for the base concrete situation


    def __init__(self, seed_boolean, seedz):  

        # Generate the ego_veh object upon initialization of the int sit framework
        # this class will give the carla transforms of ego vehicle start and goal locations, based on which intersection leg
        # is selected as the START location of ego and then based upon the start, we will further select which
        # GOAL location to choose and we will query this ego_veh object to get those carla transform of the goal location
        self.ego_veh = EgoVehicle()
        self.other_veh = OtherVehicle()
        self.env_conditions = EnvironmentalConditions()

        self.initialize_situation_dictionaries()

        if seed_boolean:
            seed(seedz)
            np.random.seed(seedz)

    # approach can either be "random" or "sitcov"
    def start_sit_config_gen(self, approach, activate_env_cond_generation=False):  #Use this method again and again with each repetition of the scenario class

        # shifting this to the end of this method so that I can generate the same random situations using the seeds I have. But first leme debug this
        self.env_conditions.generate_environmental_conditions(approach, activate_env_cond_generation);

        self.cloudiness_key = self.env_conditions.cloudiness_key
        self.precipitation_key = self.env_conditions.precipitation_key
        self.precipitation_deposits_key = self.env_conditions.precipitation_deposits_key
        self.wind_intensity_key = self.env_conditions.wind_intensity_key
        self.fog_density_key = self.env_conditions.fog_density_key
        self.fog_distance_key = self.env_conditions.fog_distance_key
        self.wetness_key = self.env_conditions.wetness_key
        self.fog_falloff_key = self.env_conditions.fog_falloff_key
        self.friction_key = self.env_conditions.friction_key

        if approach == "random":


            # ******** this line only written to run the file, or it throws an error at the moment.
            #self.ego_veh.select_ego_start_loc("left")  # just testing. Comment out when not debugging.
            #pass

            key_ego_start_index_list = []
            
            for key in self.ego_start_count_plus_other_dict.keys():

                #print(key)

                #sys.exit("off")


                #if key == "key_ego_start_left_count" or "key_ego_start_base_count" or "key_ego_start_right_count":
                if key == "key_ego_start_left_count" or key == "key_ego_start_base_count" or key == "key_ego_start_right_count":
                
                #if key == "key_ego_start_left_count":  # yes!

                    key_ego_start_index_list.append(key)
                    #print("ego start keys are as follows :3",key_ego_start_index_list)  # Works!!
                    
            
            # we have a dictionary of keys of start locations and we will randomly select one of the start locations

            #sys.exit("off")

            #seed(1)  # seed works. always generates same random numbers. great for re-creating situations.
            #key_index = randint(0,2) # works, all three conditions appear
            #key_index = randint(0,(len(key_ego_start_index_list)-1)) # works, all three conditions appear

            if len(key_ego_start_index_list) == 1:
                key_index = 0  # only one value to choose from
            elif len(key_ego_start_index_list) > 1:
                key_index = randint(0,len(key_ego_start_index_list)-1) # works, all three conditions appear
            else:
                raise ValueError("Unexpected value of key_index")

            key_ego_start_loc = key_ego_start_index_list[key_index]

            print("following key was selected :3", key_ego_start_loc)

            if key_ego_start_loc == "key_ego_start_left_count":
                # Increment this situation by one since it has been selected for situation generation. Do this for the rest of the dictionaries as well, to get multi-level situation coverage. 
                self.ego_start_count_plus_other_dict["key_ego_start_left_count"] = self.ego_start_count_plus_other_dict["key_ego_start_left_count"] + 1
                print("key_ego_start_left_count", self.ego_start_count_plus_other_dict["key_ego_start_left_count"])
                # Here we are setting the start location of ego veh object to the left. Need to pass this start location value to the configuation file as well

                self.ego_veh.select_ego_start_loc("left")  # ego_start_carla_transform_dict is set to left int leg transform                
                # This gives us the whole transform for the start location of ego vehicle
                #self.ego_veh.ego_start_carla_transform_dict["key_transform"]

                # So for a certain start locaiton, we have some triggers to set that depend upon that start location of the ego veh, they are as follows:

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("left")

                # initialize startothertrigger (for ego veh) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("left")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_left_start_goal_count_dict.keys():
                    if key == "key_ego_goal_base_count" or key == "key_ego_goal_right_count":
                        key_ego_goal_index_list.append(key)

                
                if len(key_ego_goal_index_list) == 1:
                    key_index = 0  # only one value to choose from
                elif len(key_ego_goal_index_list) > 1:
                    key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear
                else:
                    raise ValueError("Unexpected value of key_index")


                #key_index = randint(0,(len(key_ego_goal_index_list)-1))  # len(key_ego_goal_index_list) is 2, but we can not make the range for randint to be 0 - 2. Have to add a minus 1 with that in all cases.
                key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_base_count":
                    # will add counters here later
                    self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] = self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] + 1 

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("base")

                    list_of_ego_interaction_dicts = self.ego_left_start_goal_count_dict["key_ego_goal_base_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")
                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c4":  # conflict point tells us which situation dictionary was selected

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVL_GBAVxSOVR_GBOV":
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVL_GBAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVL_GBAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVL_GBAVxSOVR_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                    
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                elif key_ego_goal_loc == "key_ego_goal_right_count":
                    self.ego_left_start_goal_count_dict["key_ego_goal_right_count"] = self.ego_left_start_goal_count_dict["key_ego_goal_right_count"] + 1

                    # Set ego veh goalloc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("right")

                    list_of_ego_interaction_dicts = self.ego_left_start_goal_count_dict["key_ego_goal_right_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c4":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            #if key == "key_SAVL_GBAVxSOVR_GBOV":  # only 1 here as well
                            # ************* In scenario_runnerZ5 Big mistake here above, caught by raise ValueError! :3 ************
                            #if key == "key_SAVL_GRAVxSOVR_GBOV":  # only 1 here as well

                            if key == "key_SAVL_GRAVxSOVR_GBOV":  # only 1 here as well
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        #if key_ego_other_veh_interaction == "key_SAVL_GBAVxSOVR_GBOV":  # mistake in scenario_runnerZ5:3
                        if key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVL_GRAVxSOVR_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")
                         
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c1":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVL_GRAVxSOVB_GROV" or key == "key_SAVL_GRAVxSOVB_GLOV":  # only 1 here as well
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVB_GROV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVB_GROV"] = ego_interaction_dict["key_SAVL_GRAVxSOVB_GROV"] + 1 
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base")
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        elif key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVL_GRAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")  # added for lane problems with goal locations
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected")    

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")


            elif key_ego_start_loc == "key_ego_start_base_count":
                self.ego_start_count_plus_other_dict["key_ego_start_base_count"] = self.ego_start_count_plus_other_dict["key_ego_start_base_count"] + 1
                print("key_ego_start_base_count", self.ego_start_count_plus_other_dict["key_ego_start_base_count"])


                self.ego_veh.select_ego_start_loc("base")

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("base")

                # initialize startothertrigger (for ego) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("base")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_base_start_goal_count_dict.keys():
                    if key == "key_ego_goal_left_count" or key == "key_ego_goal_right_count":
                        key_ego_goal_index_list.append(key)

                if len(key_ego_goal_index_list) == 1:
                    key_index = 0  # only one value to choose from and it is at index 0 :3
                elif len(key_ego_goal_index_list) > 1:
                    key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear
                else:
                    raise ValueError("Unexpected value of key_index")

                #key_index = randint(0,(len(key_ego_goal_index_list)-1))
                key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_left_count":
                    self.ego_base_start_goal_count_dict["key_ego_goal_left_count"] = self.ego_base_start_goal_count_dict["key_ego_goal_left_count"] + 1
                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("left")

                    list_of_ego_interaction_dicts = self.ego_base_start_goal_count_dict["key_ego_goal_left_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")


                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []
                    
                    if ego_interaction_dict["key_conflict_point"] == "c1":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GLAV_SOVL_GROV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        
                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVB_GLAV_SOVL_GROV":
                            ego_interaction_dict["key_SAVB_GLAV_SOVL_GROV"] = ego_interaction_dict["key_SAVB_GLAV_SOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GLAV_SOVR_GLOV" or key == "key_SAVB_GLAVxSOVR_GBOV":  # 2 here
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVB_GLAV_SOVR_GLOV":
                            ego_interaction_dict["key_SAVB_GLAV_SOVR_GLOV"] = ego_interaction_dict["key_SAVB_GLAV_SOVR_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        elif key_ego_other_veh_interaction == "key_SAVB_GLAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVB_GLAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVB_GLAVxSOVR_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")


                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected") 

                elif key_ego_goal_loc == "key_ego_goal_right_count":
                    self.ego_base_start_goal_count_dict["key_ego_goal_right_count"] = self.ego_base_start_goal_count_dict["key_ego_goal_right_count"] + 1
                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("right")

                    list_of_ego_interaction_dicts = self.ego_base_start_goal_count_dict["key_ego_goal_right_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c1":  # only 1

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GRAV_SOVL_GROV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        
                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]
                                    
                        if key_ego_other_veh_interaction == "key_SAVB_GRAV_SOVL_GROV":
                            ego_interaction_dict["key_SAVB_GRAV_SOVL_GROV"] = ego_interaction_dict["key_SAVB_GRAV_SOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")        


            elif key_ego_start_loc == "key_ego_start_right_count":
                self.ego_start_count_plus_other_dict["key_ego_start_right_count"] = self.ego_start_count_plus_other_dict["key_ego_start_right_count"] + 1    
                print("key_ego_start_right_count", self.ego_start_count_plus_other_dict["key_ego_start_right_count"])
                
                self.ego_veh.select_ego_start_loc("right")

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("right")

                # initialize startothertrigger (for ego) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("right")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_right_start_goal_count_dict.keys():
                    if key == "key_ego_goal_left_count" or key == "key_ego_goal_base_count":
                        key_ego_goal_index_list.append(key)

                if len(key_ego_goal_index_list) == 1:
                    key_index = 0  # only one value to choose from and it is at index 0 :3
                elif len(key_ego_goal_index_list) > 1:
                    key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear
                else:
                    raise ValueError("Unexpected value of key_index")

                #key_index = randint(0,(len(key_ego_goal_index_list)-1))
                key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_left_count":
                    self.ego_right_start_goal_count_dict["key_ego_goal_left_count"] = self.ego_right_start_goal_count_dict["key_ego_goal_left_count"] + 1
                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("left")

                    list_of_ego_interaction_dicts = self.ego_right_start_goal_count_dict["key_ego_goal_left_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GLAVxSOVB_GLOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GLAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVR_GLAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVR_GLAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")  # added

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                elif key_ego_goal_loc == "key_ego_goal_base_count":

                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("base")

                    list_of_ego_interaction_dicts = self.ego_right_start_goal_count_dict["key_ego_goal_base_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3
                    elif len(list_of_ego_interaction_dicts) > 1:
                        key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GBAVxSOVB_GLOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GBAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVR_GBAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVR_GBAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c4": 

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GBAVxSOVL_GROV" or key == "key_SAVR_GBAV_SOVL_GBOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        
                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GBAVxSOVL_GROV":
                            ego_interaction_dict["key_SAVR_GBAVxSOVL_GROV"] = ego_interaction_dict["key_SAVR_GBAVxSOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        elif key_ego_other_veh_interaction == "key_SAVR_GBAV_SOVL_GBOV":
                            ego_interaction_dict["key_SAVR_GBAV_SOVL_GBOV"] = ego_interaction_dict["key_SAVR_GBAV_SOVL_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")

            else:

                print("Please select appropriate starting location")
                #return 0
                raise ValueError("Unexpected value for appropriate starting location")

        elif approach == "sitcov":

            key_ego_start_index_list = []
                        
            for key, value in self.ego_start_count_plus_other_dict.items():

                #print(key)

                #sys.exit("off")


                #if key == "key_ego_start_left_count" or "key_ego_start_base_count" or "key_ego_start_right_count":
                if key == "key_ego_start_left_count" or key == "key_ego_start_base_count" or key == "key_ego_start_right_count":
                
                #if key == "key_ego_start_left_count":  # yes!

                    key_ego_start_index_list.append(key)
                    #key_ego_start_count_list.append(value)  # appending the count values of the respective keys in here
                    #print("ego start keys are as follows :3",key_ego_start_index_list)  # Works!!
                    
            # make a new dictionary out of the larger dictionary elf.ego_start_count_plus_other_dict, so that we are only dealing with the counts of situations and not the further situations that arise from this larger dict
            ego_start_counts_only_dict = {keys:self.ego_start_count_plus_other_dict[keys] for keys in  key_ego_start_index_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary

            if len(key_ego_start_index_list) == 1:
                #key_index = 0  # only one value to choose from
                #key_ego_start_loc = key_ego_start_index_list[key_index]
                raise ValueError("Unexpected value of key_index")  # as its len is not 1

            elif len(key_ego_start_index_list) > 1:
                #key_index = randint(0,len(key_ego_start_index_list)-1) # works, all three conditions appear

                key_ego_start_loc = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_start_counts_only_dict)

            else:
                raise ValueError("Unexpected value of key_index")

            #key_ego_start_loc = key_ego_start_index_list[key_index]

            print("following key was selected :3", key_ego_start_loc)

            if key_ego_start_loc == "key_ego_start_left_count":
                # Increment this situation by one since it has been selected for situation generation. Do this for the rest of the dictionaries as well, to get multi-level situation coverage. 
                self.ego_start_count_plus_other_dict["key_ego_start_left_count"] = self.ego_start_count_plus_other_dict["key_ego_start_left_count"] + 1
                print("key_ego_start_left_count", self.ego_start_count_plus_other_dict["key_ego_start_left_count"])
                # Here we are setting the start location of ego veh object to the left. Need to pass this start location value to the configuation file as well

                self.ego_veh.select_ego_start_loc("left")  # ego_start_carla_transform_dict is set to left int leg transform                
                # This gives us the whole transform for the start location of ego vehicle
                #self.ego_veh.ego_start_carla_transform_dict["key_transform"]

                # So for a certain start locaiton, we have some triggers to set that depend upon that start location of the ego veh, they are as follows:

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("left")

                # initialize startothertrigger (for ego veh) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("left")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_left_start_goal_count_dict.keys():
                    if key == "key_ego_goal_base_count" or key == "key_ego_goal_right_count":
                        key_ego_goal_index_list.append(key)

                ego_goal_counts_only_dict = {keys:self.ego_left_start_goal_count_dict[keys] for keys in  key_ego_goal_index_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary
                
                if len(key_ego_goal_index_list) == 1:
                    #key_index = 0  # only one value to choose from
                    #key_ego_goal_loc = key_ego_goal_index_list[key_index]
                    raise ValueError("Unexpected value of key_index")


                elif len(key_ego_goal_index_list) > 1:
                    #key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear

                    key_ego_goal_loc = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_goal_counts_only_dict)
                    # Will increment the counter for base goal location by 3 each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3
                else:
                    raise ValueError("Unexpected value of key_index")




                #key_index = randint(0,(len(key_ego_goal_index_list)-1))  # len(key_ego_goal_index_list) is 2, but we can not make the range for randint to be 0 - 2. Have to add a minus 1 with that in all cases.
                #key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_base_count":
                    # will add counters here later
                    


                    #self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] = self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] + 1 
                    self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] = self.ego_left_start_goal_count_dict["key_ego_goal_base_count"] + 3 
                    # Will increment the counter for base goal location each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3
                        #But we will increase the counter inside this row, the discrete situation by 1 only as we want to have correct statistics at the end which we can also have by incrementing by 3, we'd only have to divide the total counts by 3 then!



                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("base")

                    list_of_ego_interaction_dicts = self.ego_left_start_goal_count_dict["key_ego_goal_base_interactions_dict"]
                    # only one possible interaction here. 1 item in list.


                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from

                    # remove this here below as only one possible interaction here. 1 item in list.
                    
                    #elif len(list_of_ego_interaction_dicts) > 1:
                    #    key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    

                    else:
                        raise ValueError("Unexpected value of key_index")
                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c4":  # conflict point tells us which situation dictionary was selected

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVL_GBAVxSOVR_GBOV":
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from
                        
                        # remove this here as only one possible interaction here. 1 item in list.

                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        


                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVL_GBAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVL_GBAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVL_GBAVxSOVR_GBOV"] + 1  # incrementing by 1 only 
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                    
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                elif key_ego_goal_loc == "key_ego_goal_right_count":
                    self.ego_left_start_goal_count_dict["key_ego_goal_right_count"] = self.ego_left_start_goal_count_dict["key_ego_goal_right_count"] + 1
                    # increment the left start right goal by one only as this has 3 options as compared to the left start base goal location


                    # Set ego veh goalloc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("right")

                    list_of_ego_interaction_dicts = self.ego_left_start_goal_count_dict["key_ego_goal_right_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        #key_index = 0  # only one value to choose from
                        raise ValueError("Unexpected value of key_index")  # as there are 2 options for dict
                        #ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    
                    elif len(list_of_ego_interaction_dicts) > 1:

                        ego_interaction_dict = self.selection_of_lower_counts_dict_from_list_of_two(list_of_ego_interaction_dicts)

                        #key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))
                    #ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    #ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c4":  # dictionaries seperated by the conflict points

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            #if key == "key_SAVL_GBAVxSOVR_GBOV":  # only 1 here as well
                            # ************* In scenario_runnerZ5 Big mistake here above, caught by raise ValueError! :3 ************
                            #if key == "key_SAVL_GRAVxSOVR_GBOV":  # only 1 here as well

                            if key == "key_SAVL_GRAVxSOVR_GBOV":  # only 1 here as well
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        # it's length is 1 

                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        

                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        #if key_ego_other_veh_interaction == "key_SAVL_GBAVxSOVR_GBOV":  # mistake in scenario_runnerZ5:3
                        if key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVL_GRAVxSOVR_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")
                         
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c1":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        '''
                        # randomly choosing goal loc for ego veh
                        key_ego_goal_index_list = []

                        for key in self.ego_left_start_goal_count_dict.keys():
                            if key == "key_ego_goal_base_count" or key == "key_ego_goal_right_count":
                                key_ego_goal_index_list.append(key)

                        ego_goal_counts_only_dict = {keys:self.self.ego_left_start_goal_count_dict[keys] for keys in  key_ego_goal_index_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary
                        
                        if len(key_ego_goal_index_list) == 1:
                            key_index = 0  # only one value to choose from
                            key_ego_goal_loc = key_ego_goal_index_list[key_index]
                        elif len(key_ego_goal_index_list) > 1:
                            #key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear

                            key_ego_goal_loc = low_count_to_higher_prob_converter_for_situation_coverage(ego_goal_counts_only_dict)
                            

                            # Will increment the counter for base goal location by 3 each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3
                        else:
                            raise ValueError("Unexpected value of key_index")

                        '''

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVL_GRAVxSOVB_GROV" or key == "key_SAVL_GRAVxSOVB_GLOV":  # only 1 here as well
                                key_ego_other_veh_interaction_list.append(key)


                        ego_interaction_counts_only_dict = {keys:ego_interaction_dict[keys] for keys in  key_ego_other_veh_interaction_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary

                        if len(key_ego_other_veh_interaction_list) == 1:
                            #key_index = 0  # only one value to choose from and it is at index 0 :3
                            raise ValueError("Unexpected value of key_index")  # As for this case length is greater than 1
                       
                        elif len(key_ego_other_veh_interaction_list) > 1:
                            #key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                       
                            key_ego_other_veh_interaction = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_interaction_counts_only_dict)
                            # key is returned above

                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        #key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVB_GROV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVB_GROV"] = ego_interaction_dict["key_SAVL_GRAVxSOVB_GROV"] + 1 
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base")
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        elif key_ego_other_veh_interaction == "key_SAVL_GRAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVL_GRAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVL_GRAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")  # added for lane problems with goal locations
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected")    

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")


            elif key_ego_start_loc == "key_ego_start_base_count":
                self.ego_start_count_plus_other_dict["key_ego_start_base_count"] = self.ego_start_count_plus_other_dict["key_ego_start_base_count"] + 1
                print("key_ego_start_base_count", self.ego_start_count_plus_other_dict["key_ego_start_base_count"])


                self.ego_veh.select_ego_start_loc("base")

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("base")

                # initialize startothertrigger (for ego) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("base")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_base_start_goal_count_dict.keys():
                    if key == "key_ego_goal_left_count" or key == "key_ego_goal_right_count":
                        key_ego_goal_index_list.append(key)

                ego_goal_counts_only_dict = {keys:self.ego_base_start_goal_count_dict[keys] for keys in  key_ego_goal_index_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary

                if len(key_ego_goal_index_list) == 1:
                    #key_index = 0  # only one value to choose from and it is at index 0 :3
                    raise ValueError("Unexpected value of key_index")


                elif len(key_ego_goal_index_list) > 1:
                    #key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear
                    key_ego_goal_loc = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_goal_counts_only_dict)

                

                else:
                    raise ValueError("Unexpected value of key_index")

                #key_index = randint(0,(len(key_ego_goal_index_list)-1))
                #key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_left_count":
                    self.ego_base_start_goal_count_dict["key_ego_goal_left_count"] = self.ego_base_start_goal_count_dict["key_ego_goal_left_count"] + 1
                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("left")

                    list_of_ego_interaction_dicts = self.ego_base_start_goal_count_dict["key_ego_goal_left_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        #key_index = 0  # only one value to choose from and it is at index 0 :3
                        raise ValueError("Unexpected value of key_index")

                    elif len(list_of_ego_interaction_dicts) > 1:
                        #key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                        ego_interaction_dict = self.selection_of_lower_counts_dict_from_list_of_two(list_of_ego_interaction_dicts)

                    else:
                        raise ValueError("Unexpected value of key_index")


                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    #ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []
                    
                    if ego_interaction_dict["key_conflict_point"] == "c1":  # which ever dict was chosen, checking it's conflict point now.

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GLAV_SOVL_GROV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        
                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        

                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        

                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVB_GLAV_SOVL_GROV":
                            ego_interaction_dict["key_SAVB_GLAV_SOVL_GROV"] = ego_interaction_dict["key_SAVB_GLAV_SOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GLAV_SOVR_GLOV" or key == "key_SAVB_GLAVxSOVR_GBOV":  # 2 here
                                key_ego_other_veh_interaction_list.append(key)

                        ego_interaction_counts_only_dict = {keys:ego_interaction_dict[keys] for keys in  key_ego_other_veh_interaction_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary

                        if len(key_ego_other_veh_interaction_list) == 1:
                            #key_index = 0  # only one value to choose from and it is at index 0 :3
                            raise ValueError("Unexpected value of key_index")


                        elif len(key_ego_other_veh_interaction_list) > 1:
                            #key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                            key_ego_other_veh_interaction = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_interaction_counts_only_dict)

                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        #key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVB_GLAV_SOVR_GLOV":
                            ego_interaction_dict["key_SAVB_GLAV_SOVR_GLOV"] = ego_interaction_dict["key_SAVB_GLAV_SOVR_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        elif key_ego_other_veh_interaction == "key_SAVB_GLAVxSOVR_GBOV":
                            ego_interaction_dict["key_SAVB_GLAVxSOVR_GBOV"] = ego_interaction_dict["key_SAVB_GLAVxSOVR_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("right")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")


                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected") 

                elif key_ego_goal_loc == "key_ego_goal_right_count":
                    self.ego_base_start_goal_count_dict["key_ego_goal_right_count"] = self.ego_base_start_goal_count_dict["key_ego_goal_right_count"] + 3
                    # Will increment the counter for base goal location each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3
                                           #But we will increase the counter inside this row, the discrete situation by 1 only as we want to have correct statistics at the end which we can also have by incrementing by 3, we'd only have to divide the total counts by 3 then!

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("right")

                    list_of_ego_interaction_dicts = self.ego_base_start_goal_count_dict["key_ego_goal_right_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3


                    #elif len(list_of_ego_interaction_dicts) > 1:
                    #   key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear

                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c1":  # only 1

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c1")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVB_GRAV_SOVL_GROV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        
                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]
                                    
                        if key_ego_other_veh_interaction == "key_SAVB_GRAV_SOVL_GROV":
                            ego_interaction_dict["key_SAVB_GRAV_SOVL_GROV"] = ego_interaction_dict["key_SAVB_GRAV_SOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")        


            elif key_ego_start_loc == "key_ego_start_right_count":
                self.ego_start_count_plus_other_dict["key_ego_start_right_count"] = self.ego_start_count_plus_other_dict["key_ego_start_right_count"] + 1    
                print("key_ego_start_right_count", self.ego_start_count_plus_other_dict["key_ego_start_right_count"])
                
                self.ego_veh.select_ego_start_loc("right")

                # initialize passthroughtrigger locs wih this as well. Use their dictionaries like the start location dictionary inside the ego veh object. 
                self.ego_veh.select_ego_passthroughtrigger_loc("right")

                # initialize startothertrigger (for ego) locs wih this as well
                self.ego_veh.select_ego_startothertrigger_loc("right")

                # randomly choosing goal loc for ego veh
                key_ego_goal_index_list = []

                for key in self.ego_right_start_goal_count_dict.keys():
                    if key == "key_ego_goal_left_count" or key == "key_ego_goal_base_count":
                        key_ego_goal_index_list.append(key)

                ego_goal_counts_only_dict = {keys:self.ego_right_start_goal_count_dict[keys] for keys in  key_ego_goal_index_list}  # getting the original count values from the self.ego_start_count_plus_other_dict dictionary        

                if len(key_ego_goal_index_list) == 1:
                    #key_index = 0  # only one value to choose from and it is at index 0 :3
                    raise ValueError("Unexpected value of key_index")

                elif len(key_ego_goal_index_list) > 1:
                    #key_index = randint(0,len(key_ego_goal_index_list)-1) # works, all three conditions appear
                    key_ego_goal_loc = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_goal_counts_only_dict)

                else:
                    raise ValueError("Unexpected value of key_index")

                #key_index = randint(0,(len(key_ego_goal_index_list)-1))
                #key_ego_goal_loc = key_ego_goal_index_list[key_index]

                if key_ego_goal_loc == "key_ego_goal_left_count":
                    self.ego_right_start_goal_count_dict["key_ego_goal_left_count"] = self.ego_right_start_goal_count_dict["key_ego_goal_left_count"] + 3
                    # Will increment the counter for base goal location each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3
                                                               #But we will increase the counter inside this row, the discrete situation by 1 only as we want to have correct statistics at the end which we can also have by incrementing by 3, we'd only have to divide the total counts by 3 then!


                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("left")

                    list_of_ego_interaction_dicts = self.ego_right_start_goal_count_dict["key_ego_goal_left_interactions_dict"]

                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        key_index = 0  # only one value to choose from and it is at index 0 :3


                    #elif len(list_of_ego_interaction_dicts) > 1:
                    #    key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                    

                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # only 1 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GLAVxSOVB_GLOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GLAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVR_GLAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVR_GLAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")  # added

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                elif key_ego_goal_loc == "key_ego_goal_base_count":

                    # will add counters here later

                    # Set ego veh goal loc/endconditiontrigger here
                    self.ego_veh.select_ego_endconditiontrigger_loc("base")

                    list_of_ego_interaction_dicts = self.ego_right_start_goal_count_dict["key_ego_goal_base_interactions_dict"]


                    #RANDOMLY chose the dictionary
                    
                    if len(list_of_ego_interaction_dicts) == 1:
                        #key_index = 0  # only one value to choose from and it is at index 0 :3
                        raise ValueError("Unexpected value of key_index")

                    elif len(list_of_ego_interaction_dicts) > 1:
                        #key_index = randint(0,len(list_of_ego_interaction_dicts)-1) # works, all three conditions appear
                        ego_interaction_dict = self.selection_of_lower_counts_dict_from_list_of_two(list_of_ego_interaction_dicts)

                    else:
                        raise ValueError("Unexpected value of key_index")

                    #key_index = randint(0,(len(list_of_ego_interaction_dicts)-1))

                    #ego_interaction_dict = list_of_ego_interaction_dicts[key_index]  # 2 option for this case :3
                    # use conflict point to identify which interaction is taking place

                    key_ego_other_veh_interaction_list = []

                    if ego_interaction_dict["key_conflict_point"] == "c2":

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c2")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GBAVxSOVB_GLOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        if len(key_ego_other_veh_interaction_list) == 1:
                            key_index = 0  # only one value to choose from and it is at index 0 :3
                        #elif len(key_ego_other_veh_interaction_list) > 1:
                        #    key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                        else:
                            raise ValueError("Unexpected value of key_index")


                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GBAVxSOVB_GLOV":
                            ego_interaction_dict["key_SAVR_GBAVxSOVB_GLOV"] = ego_interaction_dict["key_SAVR_GBAVxSOVB_GLOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("base_left_lane")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("left")
                        
                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")

                    elif ego_interaction_dict["key_conflict_point"] == "c4": 

                        # Set the conflictpoint/SyncArrivalLoc
                        self.select_conflictpoint_syncarrival_loc("c4")

                        for key in ego_interaction_dict.keys():
                            if key == "key_SAVR_GBAVxSOVL_GROV" or key == "key_SAVR_GBAV_SOVL_GBOV":  # only 1 here
                                key_ego_other_veh_interaction_list.append(key)

                        ego_interaction_counts_only_dict = {keys:ego_interaction_dict[keys] for keys in  key_ego_other_veh_interaction_list}


                        if len(key_ego_other_veh_interaction_list) == 1:
                            #key_index = 0  # only one value to choose from and it is at index 0 :3
                            raise ValueError("Unexpected value of key_index")
                        

                        elif len(key_ego_other_veh_interaction_list) > 1:
                            #key_index = randint(0,len(key_ego_other_veh_interaction_list)-1) # works, all three conditions appear
                            key_ego_other_veh_interaction = self.low_count_to_higher_prob_converter_for_situation_coverage(ego_interaction_counts_only_dict)

                        else:
                            raise ValueError("Unexpected value of key_index")

                        #key_index = randint(0,(len(key_ego_other_veh_interaction_list)-1))
                        #key_ego_other_veh_interaction = key_ego_other_veh_interaction_list[key_index]

                        if key_ego_other_veh_interaction == "key_SAVR_GBAVxSOVL_GROV":
                            ego_interaction_dict["key_SAVR_GBAVxSOVL_GROV"] = ego_interaction_dict["key_SAVR_GBAVxSOVL_GROV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("right")
                        
                        elif key_ego_other_veh_interaction == "key_SAVR_GBAV_SOVL_GBOV":
                            ego_interaction_dict["key_SAVR_GBAV_SOVL_GBOV"] = ego_interaction_dict["key_SAVR_GBAV_SOVL_GBOV"] + 1
                            # This is the base concrete situation
                            self.key_ego_other_veh_interaction_key = key_ego_other_veh_interaction

                            # Set other vehicle start and goal locs. Set other veh triggers. Set the conflictpoint/syncloctrigger also.
                            self.other_veh.select_other_vehicle_start_loc("left")

                            # I will set the stopothertrigger to the goal location of the other vehicle but rightnow we have not setup controls to steer the othervehicle to it's goal locations.
                            # I will set up the other vehicle controls to steer it to it's goal location in later stages. Maybe use a driving agent for that, but will have to see how to use that
                            # with pytrees.
                            self.other_veh.select_other_vehicle_stopothertrigger_loc("base")

                        # no other situations possible in this case
                        else:
                            raise ValueError("Unexpected value of key_ego_other_veh_interaction selected")
                    else:
                        raise ValueError("Unexpected value of conflict point selected")

                else:
                    raise ValueError("Unexpected value of goal loc of ego veh selected")

            else:

                print("Please select appropriate starting location")
                #return 0
                raise ValueError("Unexpected value for appropriate starting location")

        else:
            print("Please select approprite situation generation technique")
            #return 0
            raise ValueError("Unexpected value approprite situation generation technique")

    def initialize_situation_dictionaries(self):

        # This is the main dictionay that will hold all the further nested dictionaries 
        self.ego_start_count_plus_other_dict = dict()
        
        # We are starting with minimum granularity, i.e., 1 bin per category only.

        
        # *********************************************** LEFT ego AV Start loc *************************************
        ##########################################################################################################
        ##########################################################################################################
        ##########################################################################################################
        
        # Conflict point dictionaries for this ego AV when it starts at self.ego_left_start_goal_count_dict
        self.c1_AV_L_R_dict = {"key_SAVL_GRAVxSOVB_GROV": 0, "key_SAVL_GRAVxSOVB_GLOV": 0, 
        "key_conflict_point": "c1"}  # maybe put the location of c1 at the end of this array so you can access it by -1 index. WRONG! Can't use -1 for dictionaries xP, dict is unordered data. 
        self.c4_AV_L_R_dict = {"key_SAVL_GRAVxSOVR_GBOV": 0, "key_conflict_point": "c4"}
        


        self.c4_AV_L_B_dict = {"key_SAVL_GBAVxSOVR_GBOV": 0, "key_conflict_point": "c4"}



        # Dictionary when ego starts at left intersection leg, how many times did it go to the following possible goal locations   
        # Look at the all possible situations table mentioned in the word document to better understand this
        #SAVL Goal loc of AV (look at word doc for this terminology) GBAVL, GRAVL
        self.ego_left_start_goal_count_dict = {"key_ego_goal_base_count": 0, 
        "key_ego_goal_base_interactions_dict": [self.c4_AV_L_B_dict],  # making this dictionary a list, so we can iterate through it, in the case there are multiple dictionaries as seen in the following key values of this dictionary
        "key_ego_goal_right_count": 0, "key_ego_goal_right_interactions_dict": [self.c1_AV_L_R_dict, self.c4_AV_L_R_dict]}  # array of dictionaries
        
        # Will increment the counter for base goal location each time as the right goal location has 3 combinations to go to whereas the base goal location has only one combo. So to balance it out we have to increment its counter by 3 :3 

        

        
        # *********************************************** RIGHT ego AV Start loc *************************************
        ##########################################################################################################
        ##########################################################################################################
        ##########################################################################################################
        
        # Conflict point dictionaries for this ego AV when it starts at self.ego_right_start_goal_count_dict
        self.c2_AV_R_L_dict = {"key_SAVR_GLAVxSOVB_GLOV": 0, "key_conflict_point": "c2"}
        self.c2_AV_R_B_dict = {"key_SAVR_GBAVxSOVB_GLOV": 0,"key_conflict_point": "c2"}
        self.c4_AV_R_B_dict = {"key_SAVR_GBAVxSOVL_GROV": 0,"key_SAVR_GBAV_SOVL_GBOV": 0,
        "key_conflict_point": "c4"}


        # Dictionary when ego starts at right intersection leg, how many times did it go to the following possible goal locations
        # SAVR Goal loc of AV GLAVR, GBAVR
        self.ego_right_start_goal_count_dict = {"key_ego_goal_left_count": 0, 
        "key_ego_goal_left_interactions_dict": [self.c2_AV_R_L_dict],"key_ego_goal_base_count": 0, 
        "key_ego_goal_base_interactions_dict": [self.c2_AV_R_B_dict, self.c4_AV_R_B_dict]}




       
        # *********************************************** BASE ego AV Start loc *************************************
        ##########################################################################################################
        ##########################################################################################################
        ##########################################################################################################
        
        # Conflict point dictionaries for this ego AV when it starts at self.ego_base_start_goal_count_dict
        self.c1_AV_B_R_dict = {"key_SAVB_GRAV_SOVL_GROV": 0, "key_conflict_point": "c1"} 
        self.c1_AV_B_L_dict = {"key_SAVB_GLAV_SOVL_GROV": 0,"key_conflict_point": "c1"}
        self.c2_AV_B_L_dict = {"key_SAVB_GLAV_SOVR_GLOV": 0, "key_SAVB_GLAVxSOVR_GBOV": 0, "key_conflict_point": "c2"}


        # Dictionary when ego starts at base intersection leg, how many times did it go to the following possible goal locations
        # SAVB Goal loc of AV GLAVB, GRAVB
        self.ego_base_start_goal_count_dict = {"key_ego_goal_left_count": 0, 
        "key_ego_goal_left_interactions_dict": [self.c1_AV_B_L_dict, self.c2_AV_B_L_dict],
        "key_ego_goal_right_count": 0, "key_ego_goal_right_interactions_dict": [self.c1_AV_B_R_dict]}





        # MAIN dictionary intialization. Put this at the end.

        # Sow if we want to add more "bins" to start locations, we make the value side of the keys, an array. Like "key_ego_start_left": [0, 1, 2, 3, 4]. So value 0 corresponds
        # to original left start ego loc, then other values can be like +-5 meters or something. We can even change initial velocities.
        # I will be making bins of these vales in the later stages.
        # Start loc of AV
        self.ego_start_count_plus_other_dict = {"key_ego_start_left_count": 0, 
        "key_ego_left_start_goal_count_dict": self.ego_left_start_goal_count_dict, 
        "key_ego_start_base_count":0, "key_ego_base_start_goal_count_dict": self.ego_base_start_goal_count_dict, 
        "key_ego_start_right_count":0, "key_ego_right_start_goal_count_dict": self.ego_right_start_goal_count_dict}

        # Based on the conflict point selection, I wil have my other vehicle starting and goal locations (which I can make multiple bins of as well)


    def select_conflictpoint_syncarrival_loc(self, conflict_point="c1"):

        self.conflictpoint_syncarrival_loc_dict = dict()

        if conflict_point=="c1":

            self.conflictpoint_syncarrival_loc_x = -74.63
            self.conflictpoint_syncarrival_loc_y = -136.34

            self.conflictpoint_syncarrival_loc_dict = {"key_conflictpoint_syncarrival_loc_x": self.conflictpoint_syncarrival_loc_x,
            "key_conflictpoint_syncarrival_loc_y": self.conflictpoint_syncarrival_loc_y}

        elif conflict_point=="c2":

            self.conflictpoint_syncarrival_loc_x = -74.63
            self.conflictpoint_syncarrival_loc_y = -139.28

            self.conflictpoint_syncarrival_loc_dict = {"key_conflictpoint_syncarrival_loc_x": self.conflictpoint_syncarrival_loc_x,
            "key_conflictpoint_syncarrival_loc_y": self.conflictpoint_syncarrival_loc_y}

        elif conflict_point=="c3":

            self.conflictpoint_syncarrival_loc_x = -89.00
            self.conflictpoint_syncarrival_loc_y = -140

            self.conflictpoint_syncarrival_loc_dict = {"key_conflictpoint_syncarrival_loc_x": self.conflictpoint_syncarrival_loc_x,
            "key_conflictpoint_syncarrival_loc_y": self.conflictpoint_syncarrival_loc_y}

        elif conflict_point=="c4":

            self.conflictpoint_syncarrival_loc_x = -89.00
            self.conflictpoint_syncarrival_loc_y = -136.5

            self.conflictpoint_syncarrival_loc_dict = {"key_conflictpoint_syncarrival_loc_x": self.conflictpoint_syncarrival_loc_x,
            "key_conflictpoint_syncarrival_loc_y": self.conflictpoint_syncarrival_loc_y}

        else:
            raise ValueError("Unexpected value for conflict_point")

    def low_count_to_higher_prob_converter_for_situation_coverage(self, dict_with_counts):
        #pass
        
        keys_list = []
        counts_list = []
        
        for key, value in dict_with_counts.items():
            
            keys_list.append(key)
            counts_list.append(value)
            
        counts_np_array = np.array(counts_list)
        
        m = softmax(counts_np_array)
        #invert the probabilites i.e., p2 = 1 - p
        m_inv = 1 - m
        sum_invP = m_inv.sum()  # sum of inverted probabilities
        normalized_invP = m_inv/sum_invP  #output normalized inverted probabilities         
        
        # return choice_list
        # Weighted random choice
        Choice = np.random.choice(keys_list, size=1, p=normalized_invP)  # return one key from the dictionary based on higher probability of selection of low counts 

        # Choice is coming out to be an ndarray so I will us Choice.item() which returns a single item in its original data type
        #return Choice
        return Choice.item()


        # Add condition if len(keys_list) or len(values_list) is 1 then don't do the random selection stuff. Just do it. 

    def selection_of_lower_counts_dict_from_list_of_two_beta(self, dicts_list_of_two): # not used anywhere

        # Dicts list has two dictionaries. We know this information initially.

        # Disecting each dict in the dicts_list to sum up their total counts

        dict1 = dicts_list_of_two[0]
        dict2 = dicts_list_of_two[1]

        list_total_counts_for_two_dicts = []
        counter_dict1 = 0
        counter_dict2 = 0

        for key, value in dict1.items():

            if not isinstance(value, str): # filtering out the counter values
                counter_dict1 = value + counter_dict1  # adding all the counter values in this dict

        for key, value in dict2.items():

            if not isinstance(value, str): # filtering out the counter values
                counter_dict2 = value + counter_dict2 # adding all the counter values in this dict

        list_total_counts_for_two_dicts = [counter_dict1, counter_dict2]

        counts_np_array = np.array(list_total_counts_for_two_dicts)
        
        m = softmax(counts_np_array)
        #invert the probabilites i.e., p2 = 1 - p
        m_inv = 1 - m
        sum_invP = m_inv.sum()  # sum of inverted probabilities
        normalized_invP = m_inv/sum_invP  #output normalized inverted probabilities 

        # Weighted random choice
        Choice = np.random.choice(list_total_counts_for_two_dicts, size=1, p=normalized_invP)  # return one key from the dictionary based on higher probability of selection of low counts 

        # returning dict with lower total counts (it would have higher probability to get returned :3)

        if Choice == counter_dict1:

            return dict1

        elif Choice == counter_dict2:

            return dict2

        else:
            raise ValueError("Something is wrong here")

    def selection_of_lower_counts_dict_from_list_of_two(self, dicts_list_of_two):

        # Dicts list has two dictionaries. We know this information initially.

        # Disecting each dict in the dicts_list to sum up their total counts

        dict1 = dicts_list_of_two[0]
        dict2 = dicts_list_of_two[1]

        counter_list_dict1 = []  # two possible counters in each dictionary

        counter_list_dict2 = []

        flag_counter_dict1 = False  # This flag will become true
        flag_counter_dict2 = False


        for key, value in dict1.items():

            if not isinstance(value, str): # filtering out the counter values
                counter_list_dict1.append(value)  # adding all the counter values in this dict

            if len(counter_list_dict1) > 1:
                flag_counter_dict1 = True  # i.e., dict1 has two counters in it i.e., two possible situations


        for key, value in dict2.items():

            if not isinstance(value, str): # filtering out the counter values
                counter_list_dict2.append(value)  # adding all the counter values in this dict

            if len(counter_list_dict2) > 1:
                flag_counter_dict2 = True  # i.e., dict2 has two counters in it i.e., two possible situations

        if flag_counter_dict1 == True and flag_counter_dict2 == True:
            raise ValueError("Something is wrong here, these two dicts can't have two counters at the same time")

        if flag_counter_dict1 == True:
            counts_np_array = np.array([counter_list_dict1[0], counter_list_dict1[1], counter_list_dict2[0]])
        elif flag_counter_dict2 == True:
            counts_np_array = np.array([counter_list_dict1[0], counter_list_dict2[0], counter_list_dict2[1]])
        else:
            raise ValueError("Something is wrong here")

        #list_total_counts_for_two_dicts = [counter_dict1, counter_dict2]

        #counts_np_array = np.array(list_total_counts_for_two_dicts)
        
        m = softmax(counts_np_array)
        #invert the probabilites i.e., p2 = 1 - p
        m_inv = 1 - m
        sum_invP = m_inv.sum()  # sum of inverted probabilities
        normalized_invP = m_inv/sum_invP  #output normalized inverted probabilities 

        # Weighted random choice

        #Choice = np.random.choice(list_total_counts_for_two_dicts, size=1, p=normalized_invP)  # return one key from the dictionary based on higher probability of selection of low counts 
        
        Choice = np.random.choice(len(counts_np_array), size=1, p=normalized_invP)  # return one key from the dictionary based on higher probability of selection of low counts 

        # returning dict with lower total counts (it would have higher probability to get returned :3)
        if flag_counter_dict1 == True:
            if Choice == 0 or Choice == 1:
                return dict1
            elif Choice == 2:
                return dict2
            else:
                raise ValueError("something is wrong")

        elif flag_counter_dict2 == True:
            if Choice == 0:
                return dict1
            elif Choice == 1 or Choice == 2:
                return dict2
            else:
                raise ValueError("something is wrong")
        else:
            raise ValueError("something is wrong")
        



class EnvironmentalConditions:

    # default values of params
    cloudiness = 0  # 0 to 100 values; 6 bins of 20 increments
    # for these dictionaries, the key is an integer and it is the bin index. Each bin index corresponds to a list, with the first value of the list corresponding
    # to the env condition parameter value and second item in the list in the counter of that parameter being selected :3
    cloudiness_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}

    precipitation = 0  # 0 to 100 values; 6 bins of 20 increments
    precipitation_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}
    
    precipitation_deposits = 0  # 0 to 100 values; 6 bins of 20 increments
    precipitation_deposits_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}

    wind_intensity = 0.35  # 0 to 100 values; 6 bins of 20 increments
    wind_intensity_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}

    # Won't be using this in the start :3 #####********************************************************** imp
    sun_azimuth_angle = 0.0  # range from 0 - 180. 7 bins of 30 increments
    sun_azimuth_angle_dict = {0:[0.0, 0.0], 1:[30.0, 0.0], 2:[60.0, 0.0], 3:[90.0, 0.0], 4:[120.0, 0.0], 5:[150.0, 0.0], 6:[180.0, 0.0]}

    # Won't be using this in the start :3 #####********************************************************** imp
    sun_altitude_angle = 15.0  # range from 90 - -90. 7 bins of 30 increments
    sun_altitude_angle_dict = {0:[-90.0, 0.0], 1:[-60.0, 0.0], 2:[-30.0, 0.0], 3:[0.0, 0.0], 4:[30.0, 0.0], 5:[60.0, 0.0], 6:[90.0, 0.0]}

    fog_density = 0.0 # 0 to 100 values; 6 bins of 20 increments
    fog_density_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}

    fog_distance = 0.0  # 0 to infinity. Dividing it into 6 bins from 0 - 100
    fog_distance_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}
    
    wetness = 0.0 # 0 to 100 values; 6 bins of 20 increments
    wetness_dict = {0:[0.0, 0.0], 1:[20.0, 0.0], 2:[40.0, 0.0], 3:[60.0, 0.0], 4:[80.0, 0.0], 5:[100.0, 0.0]}
    
    fog_falloff = 0.0  # 0 to infinity. But fog goes to the ground after value of 5. So 6 bins from 0 - 5.
    fog_falloff_dict = {0:[0.0, 0.0], 1:[1.0, 0.0], 2:[2.0, 0.0], 3:[3.0, 0.0], 4:[4.0, 0.0], 5:[5.0, 0.0]}
    
    friction = 1.0 # 0 to 1. 6 bins of 0.20 each # min fricion changed from 0 to 0.1
    friction_dict = {0:[0.1, 0.0], 1:[0.20, 0.0], 2:[0.40, 0.0], 3:[0.60, 0.0], 4:[0.80, 0.0], 5:[1.0, 0.0]}


    def __init__(self):
        pass

    def low_count_to_higher_prob_converter_for_situation_coverage(self, dict_with_counts):
        #pass
        
        keys_list = []
        counts_list = []
        
        for key, value in dict_with_counts.items():
            
            keys_list.append(key)
            counts_list.append(value)
            
        counts_np_array = np.array(counts_list)
        
        m = softmax(counts_np_array)
        #invert the probabilites i.e., p2 = 1 - p
        m_inv = 1 - m
        sum_invP = m_inv.sum()  # sum of inverted probabilities
        normalized_invP = m_inv/sum_invP  #output normalized inverted probabilities         
        
        # return choice_list
        # Weighted random choice
        Choice = np.random.choice(keys_list, size=1, p=normalized_invP)  # return one key from the dictionary based on higher probability of selection of low counts 
        # Choice is coming out to be an ndarray so I will us Choice.item() which returns a single item in its original data type
        #return Choice
        return Choice.item()


        # Add condition if len(keys_list) or len(values_list) is 1 then don't do the random selection stuff. Just do it. 

    def generate_environmental_conditions(self, approach, activate=False):

        if activate:  # if not activated, pass the default values of env conditions paramters

            # self.cloudiness_key
            # self.precipitation_key
            # self.precipitation_deposits_key
            # self.wind_intensity_key
            # self.fog_density_key
            # self.fog_distance_key 
            # self.wetness_key
            # self.fog_falloff_key
            # self.friction_key

            if approach == "random":

                key_cloudiness_dict = randint(0, 5)  # will always generate random numbers (having different random selection to each other) if randint is used in two seperate lines but for the same seed those two randint values will always be those specific values that were generated for that specific seed
                self.cloudiness_key = key_cloudiness_dict
                self.cloudiness = self.cloudiness_dict[key_cloudiness_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.cloudiness_dict[key_cloudiness_dict][1] = self.cloudiness_dict[key_cloudiness_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_precipitation_dict = randint(0, 5)
                self.precipitation_key = key_precipitation_dict
                self.precipitation = self.precipitation_dict[key_precipitation_dict][0]  # passing the first value of the list which is the parameter value
                self.precipitation_dict[key_precipitation_dict][1] = self.precipitation_dict[key_precipitation_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_precipitation_deposits_dict = randint(0, 5)
                self.precipitation_deposits_key = key_precipitation_deposits_dict
                self.precipitation_deposits = self.precipitation_deposits_dict[key_precipitation_deposits_dict][0]  # passing the first value of the list which is the parameter value
                self.precipitation_deposits_dict[key_precipitation_deposits_dict][1] = self.precipitation_deposits_dict[key_precipitation_deposits_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_wind_intensity_dict = randint(0, 5)
                self.wind_intensity_key = key_wind_intensity_dict
                self.wind_intensity = self.wind_intensity_dict[key_wind_intensity_dict][0]  # passing the first value of the list which is the parameter value
                self.wind_intensity_dict[key_wind_intensity_dict][1] = self.wind_intensity_dict[key_wind_intensity_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_fog_density_dict = randint(0, 5)
                self.fog_density_key = key_fog_density_dict
                self.fog_density = self.fog_density_dict[key_fog_density_dict][0]  # passing the first value of the list which is the parameter value
                self.fog_density_dict[key_fog_density_dict][1] = self.fog_density_dict[key_fog_density_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_fog_distance_dict = randint(0, 5)
                self.fog_distance_key = key_fog_distance_dict
                self.fog_distance = self.fog_distance_dict[key_fog_distance_dict][0]  # passing the first value of the list which is the parameter value
                self.fog_distance_dict[key_fog_distance_dict][1] = self.fog_distance_dict[key_fog_distance_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_wetness_dict = randint(0, 5)
                self.wetness_key = key_wetness_dict
                self.wetness = self.wetness_dict[key_wetness_dict][0]  # passing the first value of the list which is the parameter value
                self.wetness_dict[key_wetness_dict][1] = self.wetness_dict[key_wetness_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_fog_falloff_dict = randint(0, 5)
                self.fog_falloff_key = key_fog_falloff_dict
                self.fog_falloff = self.fog_falloff_dict[key_fog_falloff_dict][0]  # passing the first value of the list which is the parameter value
                self.fog_falloff_dict[key_fog_falloff_dict][1] = self.fog_falloff_dict[key_fog_falloff_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected

                key_friction_dict = randint(0, 5)
                self.friction_key = key_friction_dict
                self.friction = self.friction_dict[key_friction_dict][0]  # passing the first value of the list which is the parameter value
                self.friction_dict[key_friction_dict][1] = self.friction_dict[key_friction_dict][1] + 1  # Adding 1 to the counter value which is located at the second place after the parameter in the list at index 1 for this particular key that has been randomly selected



            elif approach == "sitcov":

                # self.cloudiness_key
                # self.precipitation_key
                # self.precipitation_deposits_key
                # self.wind_intensity_key
                # self.fog_density_key
                # self.fog_distance_key 
                # self.wetness_key
                # self.fog_falloff_key
                # self.friction_key
                
                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                cloudiness_counts_only_dict = {keys:self.cloudiness_dict[keys][1] for keys in self.cloudiness_dict.keys()}
                #print(len(cloudiness_counts_only_dict))
                #sys.exit("off")
                # Selecting key based on the count values
                key_cloudiness_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(cloudiness_counts_only_dict)

                self.cloudiness_key = key_cloudiness_dict

                #print(key_cloudiness_dict)
                #print(type(key_cloudiness_dict))
                #sys.exit("off")
                '''
                [1]
                <class 'numpy.ndarray'>
                off
                '''
                self.cloudiness = self.cloudiness_dict[key_cloudiness_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.cloudiness_dict[key_cloudiness_dict][1] = self.cloudiness_dict[key_cloudiness_dict][1] + 1


            
                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                precipitation_counts_only_dict = {keys:self.precipitation_dict[keys][1] for keys in self.precipitation_dict.keys()}
                # Selecting key based on the count values
                key_precipitation_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(precipitation_counts_only_dict)

                self.precipitation_key = key_precipitation_dict

                self.precipitation = self.precipitation_dict[key_precipitation_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.precipitation_dict[key_precipitation_dict][1] = self.precipitation_dict[key_precipitation_dict][1] + 1


            
                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                precipitation_deposits_counts_only_dict = {keys:self.precipitation_deposits_dict[keys][1] for keys in self.precipitation_deposits_dict.keys()}
                # Selecting key based on the count values
                key_precipitation_deposits_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(precipitation_deposits_counts_only_dict)

                self.precipitation_deposits_key = key_precipitation_deposits_dict

                self.precipitation_deposits = self.precipitation_deposits_dict[key_precipitation_deposits_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.precipitation_deposits_dict[key_precipitation_deposits_dict][1] = self.precipitation_deposits_dict[key_precipitation_deposits_dict][1] + 1


                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                wind_intensity_counts_only_dict = {keys:self.wind_intensity_dict[keys][1] for keys in self.wind_intensity_dict.keys()}
                # Selecting key based on the count values
                key_wind_intensity_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(wind_intensity_counts_only_dict)

                self.wind_intensity_key = key_wind_intensity_dict

                self.wind_intensity = self.wind_intensity_dict[key_wind_intensity_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.wind_intensity_dict[key_wind_intensity_dict][1] = self.wind_intensity_dict[key_wind_intensity_dict][1] + 1


                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                fog_density_counts_only_dict = {keys:self.fog_density_dict[keys][1] for keys in self.fog_density_dict.keys()}
                # Selecting key based on the count values
                key_fog_density_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(fog_density_counts_only_dict)

                self.fog_density_key = key_fog_density_dict

                self.fog_density = self.fog_density_dict[key_fog_density_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.fog_density_dict[key_fog_density_dict][1] = self.fog_density_dict[key_fog_density_dict][1] + 1


                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                fog_distance_counts_only_dict = {keys:self.fog_distance_dict[keys][1] for keys in self.fog_distance_dict.keys()}
                # Selecting key based on the count values
                key_fog_distance_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(fog_distance_counts_only_dict)

                self.fog_distance_key = key_fog_distance_dict

                self.fog_distance = self.fog_distance_dict[key_fog_distance_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.fog_distance_dict[key_fog_distance_dict][1] = self.fog_distance_dict[key_fog_distance_dict][1] + 1


                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                wetness_counts_only_dict = {keys:self.wetness_dict[keys][1] for keys in self.wetness_dict.keys()}
                # Selecting key based on the count values
                key_wetness_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(wetness_counts_only_dict)

                self.wetness_key = key_wetness_dict

                self.wetness = self.wetness_dict[key_wetness_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.wetness_dict[key_wetness_dict][1] = self.wetness_dict[key_wetness_dict][1] + 1

                
                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                fog_falloff_counts_only_dict = {keys:self.fog_falloff_dict[keys][1] for keys in self.fog_falloff_dict.keys()}
                # Selecting key based on the count values
                key_fog_falloff_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(fog_falloff_counts_only_dict)

                self.fog_falloff_key = key_fog_falloff_dict

                self.fog_falloff = self.fog_falloff_dict[key_fog_falloff_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.fog_falloff_dict[key_fog_falloff_dict][1] = self.fog_falloff_dict[key_fog_falloff_dict][1] + 1


                # Setting up a dictionary with the keys and ONLY counter values. Parameters values aren't needed for the sitcov based selection.
                friction_counts_only_dict = {keys:self.friction_dict[keys][1] for keys in self.friction_dict.keys()}
                # Selecting key based on the count values
                key_friction_dict = self.low_count_to_higher_prob_converter_for_situation_coverage(friction_counts_only_dict)

                self.friction_key = key_friction_dict

                self.friction = self.friction_dict[key_friction_dict][0]  # passing the first value of the list which is the parameter value
                # updating the counter of the selected env condition parameter
                self.friction_dict[key_friction_dict][1] = self.friction_dict[key_friction_dict][1] + 1


            else:
                raise ValueError("Invalid sit gen approach selected")
        else:
            pass  # default values selected



class EgoVehicle:

    # Dictionary to store the ego start location as a carla Transform and to store the start location string
    ego_start_carla_transform_dict = dict()  # structure is as follows: {"key_transform": self.start_ego_base_carla_transform, "key_location_string": start_location_string}
    ego_passthroughtrigger_dict = dict()
    ego_startothertrigger_dict = dict()

    ego_endconditiontrigger_dict = dict()  # goal location for ego vehicle
    

    #def __init__(self, start_location_string="base"):  # start_location will either be "Base", "Left", "Right" 

    def __init__(self):
        #self.ego_start_carla_transform_dict = self.select_ego_loc(start_location_string)  # self.ego_start_carla_transform_dict will be a dictionary with locaiton and the location string as it's two entries
        pass

    def select_ego_start_loc(self, start_location_string="base"):

        # Base leg of intersection EGO Veh starting waypoint location/rotation details
        self.start_ego_base_carla_loc = carla.Location(x=-74.32, y=-50, z=0.5)  
        self.start_ego_base_carla_rotation = carla.Rotation(pitch=0.0, yaw=270, roll=0.0)
        # Base leg of intersection EGO Veh starting waypoint Transform
        self.start_ego_base_carla_transform = carla.Transform(self.start_ego_base_carla_loc, self.start_ego_base_carla_rotation)


        # Left leg of intersection EGO Veh starting waypoint location/rotation details
        self.start_ego_left_carla_loc = carla.Location(x=-145.36, y=-91.61, z=0.5)
        self.start_ego_left_carla_rotation = carla.Rotation(pitch=0.0, yaw=270, roll=0.0)  # 270 yaw for left leg of intersection since the road turns left and becomes straight like Base leg
        # Left leg of intersection EGO Veh starting waypoint Transform
        self.start_ego_left_carla_transform = carla.Transform(self.start_ego_left_carla_loc, self.start_ego_left_carla_rotation)


        # Right leg of intersection EGO Veh starting waypoint location/rotation details
        #self.start_ego_right_carla_loc = carla.Location(x=10.09, y=-138.7, z=0.5)  # Right leg of intersection
        
        self.start_ego_right_carla_loc = carla.Location(x=15.09, y=-138.7, z=0.5)  # Right leg of intersection
        

        self.start_ego_right_carla_rotation = carla.Rotation(pitch=0.0, yaw=180, roll=0.0)
        # Right leg of intersection EGO Veh starting waypoint Transform
        self.start_ego_right_carla_transform = carla.Transform(self.start_ego_right_carla_loc, self.start_ego_right_carla_rotation)

        if start_location_string == "base":

            self.ego_start_carla_transform_dict = {"key_transform": self.start_ego_base_carla_transform, "key_location_string": start_location_string}  # return a dictionary
            
            # Don't need to return this dictionary. Just need to fill in the values as done in the line above.
            #return self.ego_start_carla_transform_dict
        
        elif start_location_string == "left":

            self.ego_start_carla_transform_dict = {"key_transform": self.start_ego_left_carla_transform, "key_location_string": start_location_string}  # return a dictionary

            #return self.ego_start_carla_transform_dict

        elif start_location_string == "right":

            self.ego_start_carla_transform_dict = {"key_transform": self.start_ego_right_carla_transform, "key_location_string": start_location_string}  # return a dictionary

            #return self.ego_start_carla_transform_dict

        else:

            print("Please select valid start location string for ego vehicle")
            #return 0
            raise ValueError("Unexpected value for start location string for ego vehicle")

    
    def select_ego_passthroughtrigger_loc(self, start_location_string="base"):  # INCOMPLETE


        if start_location_string == "base":

            # self.ego_passthroughtrigger_min_x = -90
            # self.ego_passthroughtrigger_max_x = -70
            # self.ego_passthroughtrigger_min_y = -124
            # self.ego_passthroughtrigger_max_y = -119

            # for base, y axis is changing. And it is getting less negative moving away from the intersection. So we will bring it closer by 3 in yaxis.

            self.ego_passthroughtrigger_min_x = -90
            self.ego_passthroughtrigger_max_x = -70
            self.ego_passthroughtrigger_min_y = -127
            self.ego_passthroughtrigger_max_y = -121


            self.ego_passthroughtrigger_dict = {"key_ego_passthroughtrigger_min_x": self.ego_passthroughtrigger_min_x, 
            "key_ego_passthroughtrigger_max_x": self.ego_passthroughtrigger_max_x, 
            "key_ego_passthroughtrigger_min_y": self.ego_passthroughtrigger_min_y, 
            "key_ego_passthroughtrigger_max_y": self.ego_passthroughtrigger_max_y}

            #return self.ego_passthroughtrigger_dict  # Don't need to return anything. Only need to set the dictionary values as required by the locaiton selected.

        elif start_location_string == "left":

            # self.ego_passthroughtrigger_min_x = -106.64
            # self.ego_passthroughtrigger_max_x = -101.64
            # self.ego_passthroughtrigger_min_y = -148.5
            # self.ego_passthroughtrigger_max_y = -128.5

            # bringing this closer to intersection

            self.ego_passthroughtrigger_min_x = -103.64
            self.ego_passthroughtrigger_max_x = -98.64
            self.ego_passthroughtrigger_min_y = -148.5
            self.ego_passthroughtrigger_max_y = -128.5
            
            


            self.ego_passthroughtrigger_dict = {"key_ego_passthroughtrigger_min_x": self.ego_passthroughtrigger_min_x, 
            "key_ego_passthroughtrigger_max_x": self.ego_passthroughtrigger_max_x, 
            "key_ego_passthroughtrigger_min_y": self.ego_passthroughtrigger_min_y, 
            "key_ego_passthroughtrigger_max_y": self.ego_passthroughtrigger_max_y}

        elif start_location_string == "right":

            # self.ego_passthroughtrigger_min_x = -63.15
            # self.ego_passthroughtrigger_max_x = -58.15
            # self.ego_passthroughtrigger_min_y = -147.5
            # self.ego_passthroughtrigger_max_y = -127.5

            self.ego_passthroughtrigger_min_x = -66.15  # x axis gettind closer to intersection by 3. As it is becoming positive moving away from intersection in this case. 
            self.ego_passthroughtrigger_max_x = -61.15
            self.ego_passthroughtrigger_min_y = -147.5
            self.ego_passthroughtrigger_max_y = -127.5
            
            

            self.ego_passthroughtrigger_dict = {"key_ego_passthroughtrigger_min_x": self.ego_passthroughtrigger_min_x, 
            "key_ego_passthroughtrigger_max_x": self.ego_passthroughtrigger_max_x, 
            "key_ego_passthroughtrigger_min_y": self.ego_passthroughtrigger_min_y, 
            "key_ego_passthroughtrigger_max_y": self.ego_passthroughtrigger_max_y}

        else:
            raise ValueError("Unexpected value for select_ego_passthroughtrigger_loc for ego vehicle")

    def select_ego_startothertrigger_loc(self, start_location_string="base"):

        if start_location_string == "base":

            self.ego_startothertrigger_min_x = -80
            self.ego_startothertrigger_max_x = -70
            self.ego_startothertrigger_min_y = -75
            self.ego_startothertrigger_max_y = -60

            self.ego_startothertrigger_dict = {"key_ego_startothertrigger_min_x": self.ego_startothertrigger_min_x, 
            "key_ego_startothertrigger_max_x": self.ego_startothertrigger_max_x, 
            "key_ego_startothertrigger_min_y": self.ego_startothertrigger_min_y, 
            "key_ego_startothertrigger_max_y": self.ego_startothertrigger_max_y}

            #return self.ego_startothertrigger_dict

        elif start_location_string == "left":

            self.ego_startothertrigger_min_x = -148
            self.ego_startothertrigger_max_x = -138
            self.ego_startothertrigger_min_y = -113.5
            self.ego_startothertrigger_max_y = -98.5

            self.ego_startothertrigger_dict = {"key_ego_startothertrigger_min_x": self.ego_startothertrigger_min_x, 
            "key_ego_startothertrigger_max_x": self.ego_startothertrigger_max_x, 
            "key_ego_startothertrigger_min_y": self.ego_startothertrigger_min_y, 
            "key_ego_startothertrigger_max_y": self.ego_startothertrigger_max_y}

        elif start_location_string == "right":

            self.ego_startothertrigger_min_x = -12.5
            self.ego_startothertrigger_max_x = 2.5
            self.ego_startothertrigger_min_y = -146
            self.ego_startothertrigger_max_y = -131.1

            self.ego_startothertrigger_dict = {"key_ego_startothertrigger_min_x": self.ego_startothertrigger_min_x, 
            "key_ego_startothertrigger_max_x": self.ego_startothertrigger_max_x, 
            "key_ego_startothertrigger_min_y": self.ego_startothertrigger_min_y, 
            "key_ego_startothertrigger_max_y": self.ego_startothertrigger_max_y}

        else:
            raise ValueError("Unexpected value for select_ego_startothertrigger_loc for ego vehicle")
    
    def select_ego_endconditiontrigger_loc(self, ego_goal_location="base"):  # THE GOAL LOCATION OF EGO 

        if ego_goal_location == "base":  # maybe set this location text in the dictionary also to pass to the veh controlling agent

            # self.ego_endconditiontrigger_min_x = -90
            # self.ego_endconditiontrigger_max_x = -70
            # self.ego_endconditiontrigger_min_y = -119.92
            # self.ego_endconditiontrigger_max_y = -105.923
            
            # Re-defining new ego_endcontiontriggers according to the new ego destinations
            self.ego_endconditiontrigger_min_x = -95.20  # delta is 20 i.e. half is 10
            self.ego_endconditiontrigger_max_x = -75.20 

            #self.ego_endconditiontrigger_min_y = -37  # delta was 14
            #self.ego_endconditiontrigger_max_y = -23 # swap the min and max to be real min and max as per sign
            
            # Re-adjusting the end condition trigger location considering that ego stops around 20 meters before, longitudinally

            self.ego_endconditiontrigger_min_y = -57  # delta was 14
            # new mid is y = -50
            self.ego_endconditiontrigger_max_y = -43 # swap the min and max to be real min and max as per sign

            # lets run this and see if the ego stops here!!!!!!!! :3 
            #Works!!

            self.ego_destination_x = -85.20
            #self.ego_destination_x = -88.68
            #self.ego_destination_x = -78.15
            
            #self.ego_destination_y = -112.92
            self.ego_destination_y = -30.0  # Extended in logitudinal direction

            self.ego_endconditiontrigger_dict = {"key_ego_endconditiontrigger_min_x": self.ego_endconditiontrigger_min_x, 
            "key_ego_endconditiontrigger_max_x": self.ego_endconditiontrigger_max_x, 
            "key_ego_endconditiontrigger_min_y": self.ego_endconditiontrigger_min_y, 
            "key_ego_endconditiontrigger_max_y": self.ego_endconditiontrigger_max_y,
            "key_ego_destination_x": self.ego_destination_x, 
            "key_ego_destination_y": self.ego_destination_y}

        elif ego_goal_location == "left":

            # self.ego_endconditiontrigger_min_x = -148.5
            # self.ego_endconditiontrigger_max_x = -128.5
            # self.ego_endconditiontrigger_min_y = -118.64
            # self.ego_endconditiontrigger_max_y = -104.64

            # self.ego_endconditiontrigger_min_x = -118.64
            # self.ego_endconditiontrigger_max_x = -104.64
            # self.ego_endconditiontrigger_min_y = -148.5
            # self.ego_endconditiontrigger_max_y = -128.5
            
            # Re-defining new ego_endcontiontriggers according to the new ego destinations

            # self.ego_endconditiontrigger_min_x = -155.97  # DELTA IS 14 i.e. half is 7
            # self.ego_endconditiontrigger_max_x = -141.97
            # self.ego_endconditiontrigger_min_y = -86.77  # delat is 20 i.e. half is 10
            # self.ego_endconditiontrigger_max_y = -66.77

            # Re-adjusting the end condition trigger location considering that ego stops around 20 meters before, longitudinally

            self.ego_endconditiontrigger_min_x = -155.77  # DELTA IS 14 i.e. half is 7
            # mid x = -148.77
            self.ego_endconditiontrigger_max_x = -141.77
            

            self.ego_endconditiontrigger_min_y = -106.91  # delat is 20 i.e. half is 10
            # mid y = -96.91
            self.ego_endconditiontrigger_max_y = -86.91

            # self.ego_destination_x = -109.86
            # self.ego_destination_y = -139.8
            self.ego_destination_x = -148.97  # extended longitudinally
            self.ego_destination_y = -76.77  # extended longitudinally

            self.ego_endconditiontrigger_dict = {"key_ego_endconditiontrigger_min_x": self.ego_endconditiontrigger_min_x, 
            "key_ego_endconditiontrigger_max_x": self.ego_endconditiontrigger_max_x, 
            "key_ego_endconditiontrigger_min_y": self.ego_endconditiontrigger_min_y, 
            "key_ego_endconditiontrigger_max_y": self.ego_endconditiontrigger_max_y,
            "key_ego_destination_x": self.ego_destination_x, 
            "key_ego_destination_y": self.ego_destination_y} 

        elif ego_goal_location == "right":

            # self.ego_endconditiontrigger_min_x = -148.17
            # self.ego_endconditiontrigger_max_x = -128.17
            # self.ego_endconditiontrigger_min_y = -60.08
            # self.ego_endconditiontrigger_max_y = -46.08

            # self.ego_endconditiontrigger_min_x = -60.08  
            # self.ego_endconditiontrigger_max_x = -46.08
            # self.ego_endconditiontrigger_min_y = -148.17
            # self.ego_endconditiontrigger_max_y = -128.17

            # Re-defining new ego_endcontiontriggers according to the new ego destinations

            # self.ego_endconditiontrigger_min_x = 23.0  # delta is 14 i.e. half is 7
            # self.ego_endconditiontrigger_max_x = 37.0
            # self.ego_endconditiontrigger_min_y = -148.17
            # self.ego_endconditiontrigger_max_y = -128.17


            # Re-adjusting the end condition trigger location considering that ego stops around 20 meters before, longitudinally

            self.ego_endconditiontrigger_min_x = 3.0  # delta is 14 i.e. half is 7
            # mid x = 10 
            self.ego_endconditiontrigger_max_x = 17.0
            

            self.ego_endconditiontrigger_min_y = -148.17
            self.ego_endconditiontrigger_max_y = -128.17

            #self.ego_destination_x = -55.16
            self.ego_destination_x = 30.0  # Extended longitudinally
            self.ego_destination_y = -135.95

            self.ego_endconditiontrigger_dict = {"key_ego_endconditiontrigger_min_x": self.ego_endconditiontrigger_min_x, 
            "key_ego_endconditiontrigger_max_x": self.ego_endconditiontrigger_max_x, 
            "key_ego_endconditiontrigger_min_y": self.ego_endconditiontrigger_min_y, 
            "key_ego_endconditiontrigger_max_y": self.ego_endconditiontrigger_max_y,
            "key_ego_destination_x": self.ego_destination_x, 
            "key_ego_destination_y": self.ego_destination_y}

        else:
            raise ValueError("Unexpected value for select_ego_endconditiontrigger_loc for ego vehicle")


class OtherVehicle:

    other_vehicle_start_carla_transform_dict = dict()  # start location for other vehicle
    other_vehicle_stopothertrigger_dict = dict()  # the goal loc for other vehicle

    def __init__(self):
        
        pass

    def select_other_vehicle_start_loc(self, start_location_string="base"):

        if start_location_string == "left": # OFFICIALLY Screwed this up xDDDDDDDDDDDDDD

            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-105, y=-136, z=0.5) 
            
            # 10 m away from intersection
            self.other_vehicle_start_loc_carla_location = carla.Location(x=-109, y=-136, z=0.5) 
            # WORRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRNGGGGGGGGGGGGGGGGGGGGGGGGGGGGG
            # self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=0.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=0.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_transform = carla.Transform(self.other_vehicle_start_loc_carla_location, self.other_vehicle_start_loc_carla_rotation)

            self.other_vehicle_start_carla_transform_dict = {"key_transform": self.other_vehicle_start_loc_carla_transform, 
            "key_location_string": start_location_string}

        elif start_location_string == "base":

            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-74.70, y=-119.55, z=0.5) 
            

            # 10 m away from intersection
            self.other_vehicle_start_loc_carla_location = carla.Location(x=-74.70, y=-114.95, z=0.5) 
            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-78.70, y=-114.95, z=0.5) # works!!
             
            # WORRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRNGGGGGGGGGGGGGGGGGGGGGGGGGGGGG
            #self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=90.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=270.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_transform = carla.Transform(self.other_vehicle_start_loc_carla_location, self.other_vehicle_start_loc_carla_rotation)

            self.other_vehicle_start_carla_transform_dict = {"key_transform": self.other_vehicle_start_loc_carla_transform, 
            "key_location_string": start_location_string}

        elif start_location_string == "base_left_lane":

            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-74.70, y=-119.55, z=0.5) 
            

            # 10 m away from intersection
            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-74.70, y=-114.95, z=0.5) 
            self.other_vehicle_start_loc_carla_location = carla.Location(x=-78.70, y=-114.95, z=0.5) # works!!
             
            # WORRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRNGGGGGGGGGGGGGGGGGGGGGGGGGGGGG
            #self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=90.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=270.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_transform = carla.Transform(self.other_vehicle_start_loc_carla_location, self.other_vehicle_start_loc_carla_rotation)

            self.other_vehicle_start_carla_transform_dict = {"key_transform": self.other_vehicle_start_loc_carla_transform, 
            "key_location_string": start_location_string}

        elif start_location_string == "right":

            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-69.68, y=-139.5, z=0.5) 
            #self.other_vehicle_start_loc_carla_location = carla.Location(x=-45.08, y=-139.5, z=0.5) 
            self.other_vehicle_start_loc_carla_location = carla.Location(x=-55.08, y=-139.5, z=0.5) 
            # WORRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRNGGGGGGGGGGGGGGGGGGGGGGGGGGGGG
            #self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=270, roll=0.0 )
            self.other_vehicle_start_loc_carla_rotation = carla.Rotation(pitch=0.0, yaw=180.0, roll=0.0 )
            self.other_vehicle_start_loc_carla_transform = carla.Transform(self.other_vehicle_start_loc_carla_location, self.other_vehicle_start_loc_carla_rotation)

            self.other_vehicle_start_carla_transform_dict = {"key_transform": self.other_vehicle_start_loc_carla_transform, 
            "key_location_string": start_location_string}

        else:

            raise ValueError("Unexpected value for select_other_vehicle_start_loc for other vehicle")


    def select_other_vehicle_stopothertrigger_loc(self, other_vehicle_goal_location="base"):  # THE GOAL LOCATION OF Other vehicle 

        # move all of them 7.5 way points ahead for the other veh automatic controller and also have to make it a valid lane and goal location for the controller so I guess I'll use the values from ego end locaiton trigger values :3 leme double check through screenshots first :3

        if other_vehicle_goal_location == "base":

            # self.other_vehicle_stopothertrigger_min_x = -91.6
            # self.other_vehicle_stopothertrigger_max_x = -81.6
            # self.other_vehicle_stopothertrigger_min_y = -104.92
            # self.other_vehicle_stopothertrigger_max_y = -94.92

            # move all of them 7.5 way points ahead for the other veh automatic controller and also have to make it a valid lane and goal location for the controller so I guess I'll use the values from ego end locaiton trigger values :3 leme double check through screenshots first :3

            self.other_vehicle_stopothertrigger_min_x = -95.20  # delta is 20 i.e. half is 10
            self.other_vehicle_stopothertrigger_max_x = -75.20

            self.other_vehicle_stopothertrigger_min_y = -57    # delta was 14
            self.other_vehicle_stopothertrigger_max_y = -43


            self.other_vehicle_destination_x = -85.20
            self.other_vehicle_destination_y = -30.0  # Extended in logitudinal direction

            # WRONG BELOW!!!
            #self.ego_endconditiontrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            
            self.other_vehicle_stopothertrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            "key_other_vehicle_stopothertrigger_max_x": self.other_vehicle_stopothertrigger_max_x, 
            "key_other_vehicle_stopothertrigger_min_y": self.other_vehicle_stopothertrigger_min_y, 
            "key_other_vehicle_stopothertrigger_max_y": self.other_vehicle_stopothertrigger_max_y,
            "key_other_vehicle_destination_x": self.other_vehicle_destination_x,
            "key_other_vehicle_destination_y": self.other_vehicle_destination_y}
            
        elif other_vehicle_goal_location == "left":

            # self.other_vehicle_stopothertrigger_min_x = -129.64
            # self.other_vehicle_stopothertrigger_max_x = -119.64
            # self.other_vehicle_stopothertrigger_min_y = -142
            # self.other_vehicle_stopothertrigger_max_y = -132


            self.other_vehicle_stopothertrigger_min_x = -155.77  # DELTA IS 14 i.e. half is 7
            # mid x = -148.77
            self.other_vehicle_stopothertrigger_max_x = -141.77


            self.other_vehicle_stopothertrigger_min_y = -106.91  # delat is 20 i.e. half is 10
            # mid y = -96.91
            self.other_vehicle_stopothertrigger_max_y = -86.91


            self.other_vehicle_destination_x = -148.97  # extended longitudinally
            self.other_vehicle_destination_y = -76.77  # extended longitudinally


            #self.ego_endconditiontrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            
            self.other_vehicle_stopothertrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            "key_other_vehicle_stopothertrigger_max_x": self.other_vehicle_stopothertrigger_max_x, 
            "key_other_vehicle_stopothertrigger_min_y": self.other_vehicle_stopothertrigger_min_y, 
            "key_other_vehicle_stopothertrigger_max_y": self.other_vehicle_stopothertrigger_max_y,
            "key_other_vehicle_destination_x": self.other_vehicle_destination_x,
            "key_other_vehicle_destination_y": self.other_vehicle_destination_y}
        
        elif other_vehicle_goal_location == "right":

            # self.other_vehicle_stopothertrigger_min_x = -45
            # self.other_vehicle_stopothertrigger_max_x = -35
            # self.other_vehicle_stopothertrigger_min_y = -140
            # self.other_vehicle_stopothertrigger_max_y = -130

            self.other_vehicle_stopothertrigger_min_x = 3.0  # delta is 14 i.e. half is 7
            # mid x = 10
            self.other_vehicle_stopothertrigger_max_x = 17.0


            self.other_vehicle_stopothertrigger_min_y = -148.17
            self.other_vehicle_stopothertrigger_max_y = -128.17



            self.other_vehicle_destination_x = 30.0  # Extended longitudinally
            self.other_vehicle_destination_y = -135.95

            #self.ego_endconditiontrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            
            self.other_vehicle_stopothertrigger_dict = {"key_other_vehicle_stopothertrigger_min_x": self.other_vehicle_stopothertrigger_min_x, 
            "key_other_vehicle_stopothertrigger_max_x": self.other_vehicle_stopothertrigger_max_x, 
            "key_other_vehicle_stopothertrigger_min_y": self.other_vehicle_stopothertrigger_min_y, 
            "key_other_vehicle_stopothertrigger_max_y": self.other_vehicle_stopothertrigger_max_y,
            "key_other_vehicle_destination_x": self.other_vehicle_destination_x,
            "key_other_vehicle_destination_y": self.other_vehicle_destination_y}


        else:
            raise ValueError("Unexpected value for other_vehicle_stopothertrigger for other vehicle")




def main():
    """
    main function
    """
    description = ("CARLA Scenario Runner: Setup, Run and Evaluate scenarios using CARLA\n"
                   "Current version: " + VERSION)

    # pylint: disable=line-too-long
    parser = argparse.ArgumentParser(description=description,
                                     formatter_class=RawTextHelpFormatter)
    parser.add_argument('-v', '--version', action='version', version='%(prog)s ' + VERSION)
    parser.add_argument('--host', default='127.0.0.1',
                        help='IP of the host server (default: localhost)')
    parser.add_argument('--port', default='2000',
                        help='TCP port to listen to (default: 2000)')
    parser.add_argument('--timeout', default="10.0",
                        help='Set the CARLA client timeout value in seconds')
    parser.add_argument('--trafficManagerPort', default='8000',
                        help='Port to use for the TrafficManager (default: 8000)')
    parser.add_argument('--trafficManagerSeed', default='0',
                        help='Seed used by the TrafficManager (default: 0)')
    parser.add_argument('--sync', action='store_true',
                        help='Forces the simulation to run synchronously')
    parser.add_argument('--list', action="store_true", help='List all supported scenarios and exit')

    parser.add_argument(
        '--scenario', help='Name of the scenario to be executed. Use the preposition \'group:\' to run all scenarios of one class, e.g. ControlLoss or FollowLeadingVehicle')
    parser.add_argument('--openscenario', help='Provide an OpenSCENARIO definition')
    parser.add_argument(
        '--route', help='Run a route as a scenario (input: (route_file,scenario_file,[route id]))', nargs='+', type=str)  # route file (routes_debug.xml) and scenario file (all_towns_traffic_scenarios1_3_4.json) and route id (0) required

    parser.add_argument(
        '--agent', help="Agent used to execute the scenario. Currently only compatible with route-based scenarios.")  # can be npc_agent.py or even a humancontroller agent that uses manual_control.py or a custon defined agent :)
    parser.add_argument('--agentConfig', type=str, help="Path to Agent's configuration file", default="")

    parser.add_argument('--output', action="store_true", help='Provide results on stdout')
    parser.add_argument('--file', action="store_true", help='Write results into a txt file')
    parser.add_argument('--junit', action="store_true", help='Write results into a junit file')
    parser.add_argument('--outputDir', default='', help='Directory for output files (default: this directory)')

    parser.add_argument('--configFile', default='', help='Provide an additional scenario configuration file (*.xml)')
    parser.add_argument('--additionalScenario', default='', help='Provide additional scenario implementations (*.py)')

    parser.add_argument('--debug', action="store_true", help='Run with debug output')
    parser.add_argument('--reloadWorld', action="store_true",
                        help='Reload the CARLA world before starting a scenario (default=True)')
    parser.add_argument('--record', type=str, default='',
                        help='Path were the files will be saved, relative to SCENARIO_RUNNER_ROOT.\nActivates the CARLA recording feature and saves to file all the criteria information.')
    parser.add_argument('--randomize', action="store_true", help='Scenario parameters are randomized')
    parser.add_argument('--repetitions', default=1, type=int, help='Number of scenario executions')
    parser.add_argument('--waitForEgo', action="store_true", help='Connect the scenario to an existing ego vehicle')
    parser.add_argument('--show_int_wps', action="store_true", help='Show intersection waypoints')
    parser.add_argument('--show_int_wp_locs', action="store_true", help='Show waypoint locations, just command works if show_int_wps is activated')
    parser.add_argument('--not_visualize', action="store_false", help='Do not show pygame window for ego veh driving')
    parser.add_argument('--Activate_IntersectionScenario_Seed', action="store_true", help='Use Intersection random num see or not')
    parser.add_argument('--IntersectionScenario_Seed', default=0, type=int,
                            help='Seed used by the IntersectionScenarios (default: 0)')
    parser.add_argument('--use_sit_cov', action="store_true", help='Do situation coverage based generation')
    parser.add_argument('--disable_env_cond_gen', action="store_false", help='Disable environemntal conditions generation')

    #parser.add_argument('--IntersectionScenario_Seed', default='0',
                            #help='Seed used by the IntersectionScenarios (default: 0)')



    arguments = parser.parse_args()
    # pylint: enable=line-too-long

    if arguments.list:
        print("Currently the following scenarios are supported:")
        print(*ScenarioConfigurationParser.get_list_of_scenarios(arguments.configFile), sep='\n')
        return 1

    if not arguments.scenario and not arguments.openscenario and not arguments.route:
        print("Please specify either a scenario or use the route mode\n\n")
        parser.print_help(sys.stdout)
        return 1

    if arguments.route and (arguments.openscenario or arguments.scenario):
        print("The route mode cannot be used together with a scenario (incl. OpenSCENARIO)'\n\n")
        parser.print_help(sys.stdout)
        return 1

    if arguments.agent and (arguments.openscenario or arguments.scenario):
        print("Agents are currently only compatible with route scenarios'\n\n")
        parser.print_help(sys.stdout)
        return 1

    if arguments.route:
        arguments.reloadWorld = True

    if arguments.agent:
        arguments.sync = True

    scenario_runner = None
    result = True
    start_time = time.time()

    try:
        scenario_runner = ScenarioRunner(arguments)
        result = scenario_runner.run()

    finally:
        if scenario_runner is not None:
            scenario_runner.destroy()
            del scenario_runner   # Saving memory by clearing class object
            print("Total time taken is --- %s seconds ---" % (time.time() - start_time))
    return not result

    


if __name__ == "__main__":
    sys.exit(main())
